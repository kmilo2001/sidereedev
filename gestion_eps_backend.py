# -*- coding: utf-8 -*-
"""
gestion_eps_backend.py
=======================
Gestion de EPS / Aseguradoras — Seccion 5.4.
Disponible para Admin y OPS (carga masiva solo Admin).

Tabla real: public.eps
  id             integer  GENERATED ALWAYS AS IDENTITY
  entidad_id     integer  NOT NULL   (FK entidad)
  codigo         varchar(30)
  nombre         varchar(200)
  nit            varchar(30)
  dv             varchar(5)
  departamento   varchar(100)
  municipio      varchar(100)
  digitado_por   varchar(200)
  correo         varchar(254)
  telefono       varchar(30)
  direccion      text
  tipo           varchar(50)   -- EPS | IPS | ARL | Aseguradora | Otro
  activo         boolean DEFAULT true
  creado_en      timestamptz DEFAULT now()
  creado_por_ops integer  (FK usuario_ops, nullable)
  actualizado_en timestamptz DEFAULT now()

Tablas auxiliares:
  public.staging_eps        — area de preparacion para carga masiva
  public.carga_masiva_lote  — cabecera de cada importacion
  public.carga_masiva_error — errores por fila

Carga masiva:
  - Soporta hasta 20 000 filas por lote (CSV o XLSX).
  - Usa INSERT...ON CONFLICT para upsert por (entidad_id, codigo).
  - Registra errores en carga_masiva_error sin detener el proceso.
  - Reporta progreso via callback opcional on_progreso(fila_actual, total).
"""
from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

# Importacion compatible standalone / sistema completo
try:
    from conexion import Conexion
    _MODO = "conexion"
except ImportError:
    try:
        from db_conexion import obtener_cursor as _get_cur
        _MODO = "db_conexion"
    except ImportError:
        _MODO = None

MAX_FILAS_LOTE = 20_000

TIPOS_VALIDOS = {"EPS", "IPS", "ARL", "Aseguradora", "Regimen Especial", "Otro"}

COLUMNAS_PLANTILLA = [
    "Codigo", "Nombre", "Tipo", "Departamento", "Municipio",
    "NIT", "DV", "Correo", "Telefono", "Direccion",
]


# ══════════════════════════════════════════════════════════════
# RESULTADO
# ══════════════════════════════════════════════════════════════

@dataclass
class Resultado:
    ok:      bool
    mensaje: str
    datos:   object = field(default=None)


# ══════════════════════════════════════════════════════════════
# HELPER DE CONEXION
# ══════════════════════════════════════════════════════════════

class _Cur:
    """Contexto de cursor dict unificado para conexion.py y db_conexion.py."""
    def __init__(self):
        self._conn = None
        self._ctx  = None

    def __enter__(self):
        if _MODO == "conexion":
            self._conn = Conexion(dict_cursor=True)
            return self._conn.__enter__().cursor()
        elif _MODO == "db_conexion":
            self._ctx = _get_cur()
            return self._ctx.__enter__()
        raise RuntimeError(
            "Sin modulo de conexion. Coloca conexion.py o db_conexion.py en el proyecto."
        )

    def __exit__(self, et, ev, tb):
        if _MODO == "conexion" and self._conn:
            self._conn.__exit__(et, ev, tb)
        return False


def _v(*keys, src: dict) -> Optional[str]:
    """Extrae el primer valor no vacio de src usando multiples claves (case-insensitive)."""
    for k in keys:
        v = src.get(k) or src.get(k.lower()) or src.get(k.upper()) or src.get(k.capitalize())
        if v is not None and str(v).strip():
            return str(v).strip()
    return None


# ══════════════════════════════════════════════════════════════
# LISTADO
# ══════════════════════════════════════════════════════════════

def listar_eps(
    entidad_id: int,
    filtro:     str  = "",
    solo_activos: bool = False,
    limite:     int  = 500,
) -> list[dict]:
    """
    Lista EPS de la entidad con filtro de busqueda en tiempo real.
    Busca por nombre, codigo, NIT, municipio o tipo.
    """
    try:
        with _Cur() as cur:
            cond   = ["e.entidad_id = %s"]
            params = [entidad_id]

            if solo_activos:
                cond.append("e.activo = TRUE")

            if filtro.strip():
                cond.append(
                    "(e.nombre      ILIKE %s "
                    " OR e.codigo   ILIKE %s "
                    " OR e.nit      ILIKE %s "
                    " OR e.municipio ILIKE %s "
                    " OR e.tipo     ILIKE %s)"
                )
                like = f"%{filtro.strip()}%"
                params += [like, like, like, like, like]

            where = " AND ".join(cond)

            cur.execute(
                f"""
                SELECT
                    e.id                                          AS eps_id,
                    e.entidad_id,
                    COALESCE(e.codigo,    '')                     AS codigo,
                    COALESCE(e.nombre,    '')                     AS nombre,
                    COALESCE(e.nit,       '')                     AS nit,
                    COALESCE(e.dv,        '')                     AS dv,
                    COALESCE(e.departamento, '')                  AS departamento,
                    COALESCE(e.municipio, '')                     AS municipio,
                    COALESCE(e.digitado_por, '')                  AS digitado_por,
                    COALESCE(e.correo,    '')                     AS correo,
                    COALESCE(e.telefono,  '')                     AS telefono,
                    COALESCE(e.direccion, '')                     AS direccion,
                    COALESCE(e.tipo,      'EPS')                  AS tipo,
                    e.activo,
                    e.creado_por_ops,
                    COALESCE(u.nombre_completo, '')               AS creado_por_ops_nombre,
                    public.eps_tiene_contrato(
                        e.entidad_id, e.id, CURRENT_DATE
                    )                                             AS tiene_contrato,
                    to_char(e.creado_en AT TIME ZONE 'America/Bogota',
                            'DD/MM/YYYY')                         AS fecha_creacion,
                    to_char(e.actualizado_en AT TIME ZONE 'America/Bogota',
                            'DD/MM/YYYY HH24:MI')                 AS ultima_actualizacion
                FROM  public.eps e
                LEFT  JOIN public.usuario_ops u ON u.id = e.creado_por_ops
                WHERE {where}
                ORDER BY e.nombre
                LIMIT  %s
                """,
                params + [limite],
            )
            return [dict(r) for r in cur.fetchall()]

    except Exception as e:
        print(f"[ERROR] listar_eps: {e}")
        return []


def obtener_eps(entidad_id: int, eps_id: int) -> Optional[dict]:
    """Retorna el detalle completo de una EPS por ID."""
    try:
        with _Cur() as cur:
            cur.execute(
                """
                SELECT
                    e.id AS eps_id, e.entidad_id,
                    COALESCE(e.codigo,       '') AS codigo,
                    COALESCE(e.nombre,       '') AS nombre,
                    COALESCE(e.nit,          '') AS nit,
                    COALESCE(e.dv,           '') AS dv,
                    COALESCE(e.departamento, '') AS departamento,
                    COALESCE(e.municipio,    '') AS municipio,
                    COALESCE(e.digitado_por, '') AS digitado_por,
                    COALESCE(e.correo,       '') AS correo,
                    COALESCE(e.telefono,     '') AS telefono,
                    COALESCE(e.direccion,    '') AS direccion,
                    COALESCE(e.tipo,      'EPS') AS tipo,
                    e.activo,
                    public.eps_tiene_contrato(e.entidad_id, e.id, CURRENT_DATE) AS tiene_contrato,
                    to_char(e.creado_en AT TIME ZONE 'America/Bogota',
                            'DD/MM/YYYY HH24:MI') AS fecha_creacion,
                    to_char(e.actualizado_en AT TIME ZONE 'America/Bogota',
                            'DD/MM/YYYY HH24:MI') AS ultima_actualizacion
                FROM  public.eps e
                WHERE e.id = %s AND e.entidad_id = %s
                """,
                (eps_id, entidad_id),
            )
            row = cur.fetchone()
        return dict(row) if row else None
    except Exception as e:
        print(f"[ERROR] obtener_eps: {e}")
        return None


# ══════════════════════════════════════════════════════════════
# CREAR / ACTUALIZAR
# ══════════════════════════════════════════════════════════════

def guardar_eps(
    entidad_id:   int,
    ops_id,
    datos:        dict,
    eps_id:       Optional[int] = None,
    nombre_admin: Optional[str] = None,
) -> Resultado:
    """
    Crea o actualiza una EPS.
    - eps_id=None  → INSERT
    - eps_id=N     → UPDATE

    datos esperados (claves flexibles Titulo/minuscula):
      codigo, nombre, tipo, departamento, municipio,
      nit, dv, correo, telefono, direccion, digitado_por
    """
    nombre   = _v("nombre",       "Nombre",       src=datos)
    codigo   = _v("codigo",       "Codigo",       src=datos)
    tipo     = _v("tipo",         "Tipo",         src=datos) or "EPS"
    dpto     = _v("departamento", "Departamento", src=datos)
    mpio     = _v("municipio",    "Municipio",    src=datos)
    nit      = _v("nit",          "NIT",          src=datos)
    dv       = _v("dv",           "DV",           src=datos)
    correo   = _v("correo",       "Correo",       src=datos)
    tel      = _v("telefono",     "Telefono",     src=datos)
    direccion = _v("direccion",   "Direccion",    src=datos)
    digitado  = (_v("digitado_por", "Digitado_por", src=datos)
                 or nombre_admin)

    if not nombre:
        return Resultado(False, "El nombre de la EPS es obligatorio.")
    if len(nombre) > 200:
        return Resultado(False, "El nombre no puede superar 200 caracteres.")

    try:
        oid = int(ops_id) if ops_id and str(ops_id).strip() not in ("", "0") else None

        with _Cur() as cur:
            if eps_id:
                # ── ACTUALIZAR ────────────────────────────────
                cur.execute(
                    "SELECT id FROM public.eps WHERE id=%s AND entidad_id=%s",
                    (eps_id, entidad_id),
                )
                if not cur.fetchone():
                    return Resultado(False, "EPS no encontrada.")

                cur.execute(
                    """
                    UPDATE public.eps SET
                        codigo        = %s,
                        nombre        = %s,
                        tipo          = %s,
                        departamento  = %s,
                        municipio     = %s,
                        nit           = %s,
                        dv            = %s,
                        correo        = %s,
                        telefono      = %s,
                        direccion     = %s,
                        digitado_por  = COALESCE(%s, digitado_por),
                        actualizado_en = now()
                    WHERE id = %s AND entidad_id = %s
                    """,
                    (codigo, nombre, tipo, dpto, mpio, nit, dv,
                     correo, tel, direccion, digitado, eps_id, entidad_id),
                )
                return Resultado(True, f"EPS '{nombre}' actualizada correctamente.",
                                 {"eps_id": eps_id, "accion": "actualizada"})

            else:
                # ── CREAR ─────────────────────────────────────
                # Verificar nombre duplicado en la misma entidad
                cur.execute(
                    "SELECT id FROM public.eps "
                    "WHERE entidad_id=%s AND LOWER(TRIM(nombre))=LOWER(%s)",
                    (entidad_id, nombre),
                )
                if cur.fetchone():
                    return Resultado(False,
                                     f"Ya existe una EPS con el nombre '{nombre}'.")

                cur.execute(
                    """
                    INSERT INTO public.eps
                        (entidad_id, codigo, nombre, tipo, departamento, municipio,
                         nit, dv, correo, telefono, direccion, digitado_por,
                         activo, creado_por_ops)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s)
                    RETURNING id
                    """,
                    (entidad_id, codigo, nombre, tipo, dpto, mpio,
                     nit, dv, correo, tel, direccion, digitado, oid),
                )
                nuevo_id = cur.fetchone()["id"]
                return Resultado(True, f"EPS '{nombre}' creada correctamente.",
                                 {"eps_id": nuevo_id, "accion": "creada"})

    except Exception as e:
        msg = str(e).split("\n")[0]
        return Resultado(False, f"Error al guardar: {msg}")


# ══════════════════════════════════════════════════════════════
# CAMBIAR ESTADO
# ══════════════════════════════════════════════════════════════

def cambiar_estado_eps(
    entidad_id: int,
    eps_id:     int,
    activo:     bool,
) -> Resultado:
    """Activa o desactiva una EPS (soft-disable)."""
    accion = "activar" if activo else "desactivar"
    try:
        with _Cur() as cur:
            cur.execute(
                "SELECT nombre FROM public.eps WHERE id=%s AND entidad_id=%s",
                (eps_id, entidad_id),
            )
            row = cur.fetchone()
            if not row:
                return Resultado(False, "EPS no encontrada.")
            nombre = row["nombre"]

            # Advertencia: al desactivar verificar si hay pacientes activos
            advertencia = ""
            if not activo:
                cur.execute(
                    "SELECT COUNT(*) AS n FROM public.paciente "
                    "WHERE eps_id=%s AND entidad_id=%s AND activo=TRUE",
                    (eps_id, entidad_id),
                )
                n = cur.fetchone()["n"]
                if n:
                    advertencia = (
                        f" Atencion: {n} paciente(s) tienen esta EPS asignada."
                    )

            cur.execute(
                "UPDATE public.eps SET activo=%s, actualizado_en=now() "
                "WHERE id=%s AND entidad_id=%s",
                (activo, eps_id, entidad_id),
            )

        lbl = "activada" if activo else "desactivada"
        return Resultado(True, f"'{nombre}' {lbl} correctamente.{advertencia}")

    except Exception as e:
        msg = str(e).split("\n")[0]
        return Resultado(False, f"Error al {accion}: {msg}")


# ══════════════════════════════════════════════════════════════
# ELIMINAR
# ══════════════════════════════════════════════════════════════

def eliminar_eps(
    entidad_id: int,
    eps_id:     int,
) -> Resultado:
    """
    Elimina fisicamente una EPS si no tiene dependencias.
    Bloquea si hay pacientes, eventos o contratos vinculados.
    """
    try:
        with _Cur() as cur:
            cur.execute(
                "SELECT nombre FROM public.eps WHERE id=%s AND entidad_id=%s",
                (eps_id, entidad_id),
            )
            row = cur.fetchone()
            if not row:
                return Resultado(False, "EPS no encontrada.")
            nombre = row["nombre"]

            # Verificar dependencias
            deps = []
            for tabla, col, lbl in [
                ("public.paciente",   "eps_id", "paciente(s)"),
                ("public.evento",     "eps_id", "evento(s)"),
                ("public.contrato_eps","eps_id", "contrato(s)"),
            ]:
                cur.execute(
                    f"SELECT COUNT(*) AS n FROM {tabla} WHERE {col}=%s",
                    (eps_id,),
                )
                n = cur.fetchone()["n"]
                if n:
                    deps.append(f"{n} {lbl}")

            if deps:
                return Resultado(
                    False,
                    f"No se puede eliminar '{nombre}': "
                    f"esta referenciada por {', '.join(deps)}. "
                    "Usa 'Desactivar' en su lugar."
                )

            cur.execute(
                "DELETE FROM public.eps WHERE id=%s AND entidad_id=%s",
                (eps_id, entidad_id),
            )

        return Resultado(True, f"EPS '{nombre}' eliminada correctamente.")

    except Exception as e:
        msg = str(e).split("\n")[0]
        return Resultado(False, f"Error al eliminar: {msg}")


# ══════════════════════════════════════════════════════════════
# CARGA MASIVA — hasta 20 000 filas
# ══════════════════════════════════════════════════════════════

def procesar_carga_masiva(
    entidad_id:   int,
    ops_id,
    ruta_archivo: str,
    nombre_admin: Optional[str] = None,
    on_progreso:  Optional[Callable[[int, int], None]] = None,
) -> Resultado:
    """
    Procesa un archivo CSV o XLSX con hasta 20 000 filas de EPS.

    Estrategia:
      - Upsert por (entidad_id, codigo): si ya existe el codigo → UPDATE,
        si no existe → INSERT.
      - Registra errores en carga_masiva_error sin detener el proceso.
      - Llama on_progreso(fila_actual, total) cada 100 filas si se provee.
      - Retorna Resultado con datos: {total, creados, actualizados, errores}.
    """
    ruta = Path(ruta_archivo)
    if not ruta.exists():
        return Resultado(False, "El archivo no existe.")

    # ── Leer filas ────────────────────────────────────────────
    filas: list[dict] = []
    try:
        ext = ruta.suffix.lower()
        if ext == ".csv":
            with open(ruta, newline="", encoding="utf-8-sig") as f:
                filas = list(csv.DictReader(f))
        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb  = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
                ws  = wb.active
                # Detectar fila de encabezados (busca "Codigo" o "Nombre" en primeras 20 filas)
                hdr_row = 1
                for row in ws.iter_rows(min_row=1, max_row=20):
                    vals = [str(c.value or "").strip().rstrip(" *") for c in row]
                    if any(v in ("Codigo", "Nombre") for v in vals):
                        hdr_row = row[0].row
                        break
                enc = [
                    str(c.value or "").strip().rstrip(" *") if c.value else None
                    for c in ws[hdr_row]
                ]
                # Saltar la fila de ayudas (justo debajo del encabezado)
                for row in ws.iter_rows(min_row=hdr_row + 2, values_only=True):
                    if all(v is None or str(v).strip() == "" for v in row):
                        continue
                    filas.append(dict(zip(enc, [str(v).strip() if v is not None else "" for v in row])))
                wb.close()
            except ImportError:
                return Resultado(False, "Para XLSX instala: pip install openpyxl")
        else:
            return Resultado(False, "Formato no soportado. Usa .csv o .xlsx")
    except Exception as e:
        return Resultado(False, f"Error al leer el archivo: {e}")

    if not filas:
        return Resultado(False, "El archivo esta vacio o no tiene datos validos.")

    total = len(filas)
    if total > MAX_FILAS_LOTE:
        return Resultado(
            False,
            f"El archivo tiene {total} filas. El maximo por lote es {MAX_FILAS_LOTE}."
        )

    # ── Crear registro de lote ────────────────────────────────
    lote_id: Optional[int] = None
    try:
        with _Cur() as cur:
            cur.execute(
                """
                INSERT INTO public.carga_masiva_lote
                    (entidad_id, tipo, nombre_archivo, total_filas, estado)
                VALUES (%s, 'eps', %s, %s, 'procesando')
                RETURNING id
                """,
                (entidad_id, ruta.name, total),
            )
            lote_id = cur.fetchone()["id"]
    except Exception:
        pass  # si falla el log de lote, igual procesamos

    # ── Procesar fila a fila ──────────────────────────────────
    creados = 0; actualizados = 0
    errores_log: list[dict] = []
    errores_bd:  list[tuple] = []  # para bulk insert en carga_masiva_error

    for idx, fila in enumerate(filas, start=2):

        def g(*keys):
            return _v(*keys, src=fila) or ""

        nombre = g("Nombre", "nombre")
        codigo = g("Codigo", "codigo").upper() or None

        if not nombre:
            e = {"fila": idx, "campo": "Nombre", "error": "Campo obligatorio"}
            errores_log.append(e)
            errores_bd.append((lote_id, idx, "Nombre", "Campo obligatorio"))
            continue

        datos = {
            "codigo":       codigo,
            "nombre":       nombre,
            "tipo":         g("Tipo",         "tipo")         or "EPS",
            "departamento": g("Departamento", "departamento"),
            "municipio":    g("Municipio",    "municipio"),
            "nit":          g("NIT",          "nit"),
            "dv":           g("DV",           "dv"),
            "correo":       g("Correo",       "correo"),
            "telefono":     g("Telefono",     "telefono"),
            "direccion":    g("Direccion",    "direccion"),
        }

        # Upsert por codigo si existe
        eps_existente: Optional[int] = None
        if codigo:
            try:
                with _Cur() as cur:
                    cur.execute(
                        "SELECT id FROM public.eps "
                        "WHERE entidad_id=%s AND codigo=%s",
                        (entidad_id, codigo),
                    )
                    row = cur.fetchone()
                    if row:
                        eps_existente = row["id"]
            except Exception:
                pass

        res = guardar_eps(entidad_id, ops_id, datos, eps_existente,
                          nombre_admin=nombre_admin)

        if res.ok:
            if (res.datos or {}).get("accion") == "actualizada":
                actualizados += 1
            else:
                creados += 1
        else:
            e = {"fila": idx, "campo": "—", "error": res.mensaje}
            errores_log.append(e)
            errores_bd.append((lote_id, idx, "—", res.mensaje))

        # Progreso cada 100 filas
        if on_progreso and idx % 100 == 0:
            on_progreso(idx - 1, total)

    # Progreso final
    if on_progreso:
        on_progreso(total, total)

    # ── Registrar errores en BD ───────────────────────────────
    if lote_id and errores_bd:
        try:
            with _Cur() as cur:
                for lote_id_, fila_n, campo, desc in errores_bd:
                    if lote_id_ is None:
                        continue
                    cur.execute(
                        "INSERT INTO public.carga_masiva_error "
                        "(lote_id, numero_fila, campo, descripcion) "
                        "VALUES (%s,%s,%s,%s)",
                        (lote_id_, fila_n, campo, desc),
                    )
        except Exception:
            pass

    # ── Cerrar lote ───────────────────────────────────────────
    if lote_id:
        try:
            with _Cur() as cur:
                cur.execute(
                    """
                    UPDATE public.carga_masiva_lote
                    SET filas_ok    = %s,
                        filas_error = %s,
                        estado      = 'completado',
                        completado_en = now()
                    WHERE id = %s
                    """,
                    (creados + actualizados, len(errores_log), lote_id),
                )
        except Exception:
            pass

    return Resultado(
        ok=True,
        mensaje=(
            f"Carga EPS completada: {creados} nuevas, "
            f"{actualizados} actualizadas, "
            f"{len(errores_log)} errores de {total} filas."
        ),
        datos={
            "total":        total,
            "creados":      creados,
            "actualizados": actualizados,
            "errores":      errores_log,
            "lote_id":      lote_id,
        },
    )


# ══════════════════════════════════════════════════════════════
# PLANTILLA EXCEL
# ══════════════════════════════════════════════════════════════

def generar_plantilla_csv(ruta: str) -> Resultado:
    """Alias que genera plantilla en xlsx aunque la ruta diga .csv."""
    ruta_xlsx = str(Path(ruta).with_suffix(".xlsx"))
    return generar_plantilla_excel(ruta_xlsx)


def generar_plantilla_excel(ruta: str) -> Resultado:
    """Genera plantilla Excel (.xlsx) con instrucciones y ejemplos."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        AZ = "2D6ADF"; AZ_C = "EEF3FF"; GR = "F5F7FA"; BL = "FFFFFF"; BO = "D1D5DB"

        def _fill(c):  return PatternFill("solid", fgColor=c)
        def _font(bold=False, color="1A1A2E", size=10, italic=False):
            return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
        def _aln(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        def _brd():
            s = Side(border_style="thin", color=BO)
            return Border(left=s, right=s, top=s, bottom=s)

        wb = Workbook(); ws = wb.active; ws.title = "Carga EPS"

        # Titulo
        ws.merge_cells("A1:J1")
        ws["A1"].value     = "Sistema Gestion Eventos — Plantilla Carga Masiva EPS"
        ws["A1"].font      = _font(bold=True, color=BL, size=13)
        ws["A1"].fill      = _fill(AZ)
        ws["A1"].alignment = _aln("left", "center")
        ws.row_dimensions[1].height = 28

        instrucciones = [
            "INSTRUCCIONES:",
            "* Campos obligatorios: Nombre.",
            "* Tipo valido: EPS, IPS, ARL, Aseguradora, Regimen Especial, Otro.",
            "* Maximo 20.000 filas por archivo.",
            "* No modificar los encabezados. Datos desde la fila 11.",
        ]
        for i, txt in enumerate(instrucciones, 2):
            ws.merge_cells(f"A{i}:J{i}")
            c = ws[f"A{i}"]
            c.value     = txt
            c.font      = _font(size=9, italic=(i > 2), bold=(i == 2), color="374151")
            c.fill      = _fill(AZ_C)
            c.alignment = _aln("left", "center", wrap=True)
            ws.row_dimensions[i].height = 14

        ws.merge_cells("A7:J7")
        ws["A7"].fill = _fill(AZ)
        ws.row_dimensions[7].height = 4

        COLS = [
            ("Codigo",       14), ("Nombre *",    36), ("Tipo",         16),
            ("Departamento", 20), ("Municipio",   20), ("NIT",          16),
            ("DV",            6), ("Correo",      30), ("Telefono",     16),
            ("Direccion",    36),
        ]
        HDR = 8
        for col, (titulo, ancho) in enumerate(COLS, 1):
            ws.column_dimensions[get_column_letter(col)].width = ancho
            c = ws.cell(row=HDR, column=col, value=titulo)
            c.font      = _font(bold=True, color=BL, size=10)
            c.fill      = _fill(AZ)
            c.alignment = _aln("center", "center")
            c.border    = _brd()
        ws.row_dimensions[HDR].height = 22

        ayudas = [
            "Cod. MSPS/SNS", "Obligatorio", "EPS/IPS/ARL...",
            "Texto", "Texto", "Solo numeros",
            "1 digito", "correo@", "Solo numeros", "Texto",
        ]
        for col, txt in enumerate(ayudas, 1):
            c = ws.cell(row=HDR + 1, column=col, value=txt)
            c.font      = _font(size=8, italic=True, color="6B7280")
            c.fill      = _fill(GR)
            c.alignment = _aln("center", "center", wrap=True)
            c.border    = _brd()
        ws.row_dimensions[HDR + 1].height = 16

        ejemplos = [
            ["EPS001", "Sanitas EPS",     "EPS",         "Cundinamarca", "Bogota",     "800251440", "5", "servicios@sanitas.com",     "6012343434", "Cra 9 No 115-06"],
            ["EPS002", "Nueva EPS",       "EPS",         "Antioquia",    "Medellin",   "900156264", "9", "servicio@nuevaeps.com",     "6043077022", "Cra 85K 46A-6"],
            ["ARL001", "Positiva ARL",    "ARL",         "Bogota",       "Bogota",     "900255549", "3", "info@positiva.com.co",       "6013207200", "Cra 7 No 14-88"],
            ["IPS001", "Clinica Ejemplo", "IPS",         "Valle",        "Cali",       "800201781", "2", "contacto@clinicaej.com",    "6023334567", "Cll 5 No 38-92"],
            ["ASG001", "Colmena Seguros", "Aseguradora", "Bogota",       "Bogota",     "830002236", "1", "info@colmena.com.co",        "6015920000", "Cll 72 No 10-03"],
        ]
        for fi, fila_ej in enumerate(ejemplos, HDR + 2):
            for col, val in enumerate(fila_ej, 1):
                c = ws.cell(row=fi, column=col, value=val)
                c.font      = _font(size=9, color="374151")
                c.fill      = _fill(BL if fi % 2 == 0 else GR)
                c.alignment = _aln("left", "center")
                c.border    = _brd()
            ws.row_dimensions[fi].height = 18

        ws.freeze_panes = ws.cell(row=HDR + 2, column=1)
        wb.save(ruta)
        return Resultado(True, f"Plantilla guardada en: {ruta}")

    except ImportError:
        return Resultado(False, "Instala openpyxl: pip install openpyxl")
    except Exception as e:
        return Resultado(False, f"Error al generar plantilla: {e}")


# ══════════════════════════════════════════════════════════════
# HELPERS STANDALONE (prueba modulo por modulo)
# ══════════════════════════════════════════════════════════════

def listar_entidades_disponibles() -> list[dict]:
    """
    Retorna todas las entidades (IPS) registradas en la BD.
    Usado en modo standalone para que el desarrollador elija
    con cual entidad_id trabajar sin necesidad de sesion.
    Retorna: [{id, nombre_entidad, nit}]
    """
    try:
        with _Cur() as cur:
            cur.execute(
                "SELECT id, nombre_entidad, nit "
                "FROM public.entidad WHERE activo = TRUE "
                "ORDER BY nombre_entidad"
            )
            return [dict(r) for r in cur.fetchall()]
    except Exception as e:
        print(f"[ERROR] listar_entidades_disponibles: {e}")
        return []


def resolver_entidad_standalone(entidad_id_hint: int = 1) -> Optional[int]:
    """
    Intenta resolver un entidad_id valido para modo standalone.

    Logica:
      1. Si entidad_id_hint existe en la BD → lo retorna directamente.
      2. Si no existe, busca la primera entidad activa disponible.
      3. Si no hay ninguna → retorna None (la UI mostrara el error).
    """
    try:
        with _Cur() as cur:
            # Intentar el hint primero
            cur.execute(
                "SELECT id FROM public.entidad WHERE id=%s AND activo=TRUE",
                (entidad_id_hint,),
            )
            if cur.fetchone():
                return entidad_id_hint

            # Fallback: primera entidad activa
            cur.execute(
                "SELECT id FROM public.entidad WHERE activo=TRUE ORDER BY id LIMIT 1"
            )
            row = cur.fetchone()
            return row["id"] if row else None

    except Exception as e:
        print(f"[ERROR] resolver_entidad_standalone: {e}")
        return None


# ══════════════════════════════════════════════════════════════
# HELPER PARA OTROS MODULOS
# ══════════════════════════════════════════════════════════════

def obtener_eps_activas(entidad_id: int) -> list[dict]:
    """
    Lista simplificada de EPS activas para poblar selectores.
    Retorna [{eps_id, codigo, nombre, tiene_contrato}].
    """
    return [
        {
            "eps_id":         r["eps_id"],
            "codigo":         r.get("codigo", ""),
            "nombre":         r.get("nombre", ""),
            "tiene_contrato": r.get("tiene_contrato", False),
        }
        for r in listar_eps(entidad_id, solo_activos=True, limite=1000)
    ]
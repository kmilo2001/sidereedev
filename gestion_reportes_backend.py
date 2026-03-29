# -*- coding: utf-8 -*-
# =============================================================================
# gestion_reportes_backend.py  --  Modulo de Reportes SIGES (v2)
#
# 4 tipos de reporte:
#   1. PRODUCCION GENERAL  -- todos los eventos del periodo
#   2. FACTURACION         -- solo eventos Terminados (estado_id=2)
#   3. CARTERA OPERATIVA   -- solo eventos Pendientes (estado_id=1)
#   4. ESTRATEGICO POR EPS -- consolidado por EPS con contrato vs sin contrato
#
# Visibilidad:
#   ops    --> filtra creado_por_ops = ops_id
#   maestro / entidad / admin --> sin filtro (ve todos)
#
# Exportacion:
#   PDF  --> reportlab (tabla + encabezado + firma)
#   XLSX --> openpyxl  (tabla + totales + formato profesional)
#
# Conexion: usa conexion.Conexion(dict_cursor=True) directamente.
# =============================================================================
from __future__ import annotations
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path

from conexion import Conexion


_ROLES_VER_TODO = {"admin", "maestro", "entidad"}

# IDs de estado en la BD
ESTADO_PENDIENTE  = 1
ESTADO_TERMINADO  = 2


# ══════════════════════════════════════════════════════════════
# RESULTADO
# ══════════════════════════════════════════════════════════════

@dataclass
class Resultado:
    ok:      bool
    mensaje: str
    datos:   object = field(default=None)


def _ops_safe(ops_id) -> int | None:
    if ops_id is None or ops_id == 0 or str(ops_id) == "":
        return None
    return int(ops_id)


def _ops_filtro(rol: str, ops_id) -> int | None:
    """ops --> filtra por ID; maestro/entidad/admin --> None (ve todos)."""
    if rol in _ROLES_VER_TODO:
        return None
    return _ops_safe(ops_id)


# ══════════════════════════════════════════════════════════════
# FECHAS PREDEFINIDAS
# ══════════════════════════════════════════════════════════════

def fechas_filtro(periodo: str) -> tuple[str | None, str | None]:
    """'hoy'|'semana'|'mes'|'todos' -> ('YYYY-MM-DD', 'YYYY-MM-DD')"""
    hoy = date.today()
    if periodo == "hoy":
        s = str(hoy); return s, s
    if periodo == "semana":
        lunes = hoy - timedelta(days=hoy.weekday())
        return str(lunes), str(hoy)
    if periodo == "mes":
        return str(hoy.replace(day=1)), str(hoy)
    return None, None


# ══════════════════════════════════════════════════════════════
# SQL BASE REUTILIZABLE
# ══════════════════════════════════════════════════════════════

def _conds_base(entidad_id: int, oid: int | None,
                fecha_desde: str | None, fecha_hasta: str | None,
                estado_id: int | None = None,
                solo_activos: bool = True) -> tuple[list, list]:
    """Construye condiciones y params reutilizables para todos los reportes."""
    conds  = ["e.entidad_id = %s"]
    params = [entidad_id]
    if solo_activos:
        conds.append("e.activo = TRUE")
    if oid is not None:
        conds.append("e.creado_por_ops = %s")
        params.append(oid)
    if estado_id is not None:
        conds.append("e.estado_id = %s")
        params.append(estado_id)
    if fecha_desde:
        conds.append("e.fecha_evento >= %s::date")
        params.append(fecha_desde)
    if fecha_hasta:
        conds.append("e.fecha_evento <= %s::date")
        params.append(fecha_hasta)
    return conds, params


_SQL_EVENTO_COLS = """
    SELECT
        e.id                                                        AS evento_id,
        CONCAT_WS(' ', p.primer_nombre, p.segundo_nombre,
                  p.primer_apellido, p.segundo_apellido)            AS nombre_paciente,
        td.abreviatura                                              AS tipo_identificacion,
        p.numero_documento                                          AS numero_identificacion,
        ep.nombre                                                   AS eps,
        ep.id                                                       AS eps_id,
        ta.nombre                                                   AS tipo_afiliacion,
        e.motivo                                                    AS motivo_evento,
        e.numero_admision,
        e.codigo_evento,
        e.numero_factura,
        e.valor,
        e.fecha_evento,
        e.tiene_contrato,
        ee.nombre                                                   AS estado_evento,
        e.estado_id,
        COALESCE(u.nombre_completo, 'Admin')                        AS registrado_por,
        (CURRENT_DATE - e.fecha_evento)                             AS dias_transcurridos
    FROM  public.evento           e
    JOIN  public.paciente         p  ON p.id  = e.paciente_id
    JOIN  public.tipo_documento   td ON td.id = p.tipo_documento_id
    JOIN  public.tipo_afiliacion  ta ON ta.id = e.tipo_afiliacion_id
    JOIN  public.estado_evento    ee ON ee.id = e.estado_id
    LEFT  JOIN public.eps              ep ON ep.id = e.eps_id
    LEFT  JOIN public.usuario_ops      u  ON u.id  = e.creado_por_ops
"""


# ══════════════════════════════════════════════════════════════
# 1. REPORTE DE PRODUCCION GENERAL
# ══════════════════════════════════════════════════════════════

def reporte_produccion(entidad_id: int, rol: str = "ops", ops_id=None,
                       fecha_desde: str | None = None,
                       fecha_hasta: str | None = None) -> list[dict]:
    """
    Todos los eventos activos del periodo.
    OPS: solo sus eventos. Maestro/Entidad/Admin: todos.
    Incluye: paciente, EPS, tipo afiliacion, numero_admision,
             codigo_evento, numero_factura, valor, fecha, ops, estado.
    """
    oid = _ops_filtro(rol, ops_id)
    conds, params = _conds_base(entidad_id, oid, fecha_desde, fecha_hasta)
    where = " AND ".join(conds)

    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        # Intentar RPC nativa primero
        try:
            cur.execute(
                """SELECT * FROM public.buscar_eventos(
                       %s::integer, %s::integer, NULL::text,
                       NULL::smallint, %s::date, %s::date,
                       FALSE, 5000::integer, 0::integer, FALSE::boolean
                   )""",
                (entidad_id, oid, fecha_desde or None, fecha_hasta or None)
            )
            rows = [dict(r) for r in cur.fetchall()]
            # Enriquecer con codigo_evento si la RPC no lo devuelve
            if rows and "codigo_evento" not in rows[0]:
                ids = [r.get("evento_id") for r in rows if r.get("evento_id")]
                if ids:
                    cur.execute(
                        "SELECT id, codigo_evento FROM public.evento"
                        " WHERE id = ANY(%s::bigint[])",
                        (ids,)
                    )
                    extras = {r["id"]: r["codigo_evento"] for r in cur.fetchall()}
                    for r in rows:
                        r["codigo_evento"] = extras.get(r.get("evento_id"), "")
            if rows:
                return rows
        except Exception:
            pass

        # Fallback SQL directo
        cur.execute(
            _SQL_EVENTO_COLS +
            f" WHERE {where}"
            "  ORDER BY e.fecha_evento DESC, e.id DESC",
            params
        )
        return [dict(r) for r in cur.fetchall()]


# ══════════════════════════════════════════════════════════════
# 2. REPORTE DE FACTURACION / EVENTOS TERMINADOS
# ══════════════════════════════════════════════════════════════

def reporte_facturacion(entidad_id: int, rol: str = "ops", ops_id=None,
                        fecha_desde: str | None = None,
                        fecha_hasta: str | None = None) -> list[dict]:
    """
    Solo eventos Terminados (estado_id=2).
    Muestra: paciente, EPS, valor, numero_factura, codigo_evento,
             numero_admision, fecha, ops responsable.
    """
    oid = _ops_filtro(rol, ops_id)
    conds, params = _conds_base(
        entidad_id, oid, fecha_desde, fecha_hasta,
        estado_id=ESTADO_TERMINADO
    )
    where = " AND ".join(conds)

    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            _SQL_EVENTO_COLS +
            f" WHERE {where}"
            "  ORDER BY e.fecha_evento DESC, e.id DESC",
            params
        )
        return [dict(r) for r in cur.fetchall()]


# ══════════════════════════════════════════════════════════════
# 3. REPORTE DE CARTERA OPERATIVA / PENDIENTES
# ══════════════════════════════════════════════════════════════

def reporte_cartera(entidad_id: int, rol: str = "ops", ops_id=None,
                    fecha_desde: str | None = None,
                    fecha_hasta: str | None = None) -> list[dict]:
    """
    Solo eventos Pendientes (estado_id=1).
    Muestra: paciente, EPS, numero_admision, dias_transcurridos, ops.
    Ordenado por dias_transcurridos DESC para priorizar los mas urgentes.
    """
    oid = _ops_filtro(rol, ops_id)
    conds, params = _conds_base(
        entidad_id, oid, fecha_desde, fecha_hasta,
        estado_id=ESTADO_PENDIENTE
    )
    where = " AND ".join(conds)

    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            _SQL_EVENTO_COLS +
            f" WHERE {where}"
            "  ORDER BY dias_transcurridos DESC, e.id DESC",
            params
        )
        return [dict(r) for r in cur.fetchall()]


# ══════════════════════════════════════════════════════════════
# 4. REPORTE ESTRATEGICO POR EPS
# ══════════════════════════════════════════════════════════════

def reporte_eps(entidad_id: int, rol: str = "ops", ops_id=None,
                fecha_desde: str | None = None,
                fecha_hasta: str | None = None) -> list[dict]:
    """
    Consolidado por EPS:
      total_eventos, total_facturado, eventos_con_contrato,
      eventos_sin_contrato, pct_participacion.
    OPS: solo sus EPS. Maestro/Entidad/Admin: consolidado general.
    """
    oid = _ops_filtro(rol, ops_id)
    conds, params = _conds_base(entidad_id, oid, fecha_desde, fecha_hasta)
    where = " AND ".join(conds)

    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            WITH base AS (
                SELECT
                    COALESCE(ep.nombre, 'Particular / Sin EPS') AS eps,
                    e.tiene_contrato,
                    e.valor,
                    e.estado_id
                FROM  public.evento           e
                LEFT  JOIN public.eps              ep ON ep.id = e.eps_id
                JOIN  public.paciente         p  ON p.id = e.paciente_id
                JOIN  public.tipo_documento   td ON td.id = p.tipo_documento_id
                JOIN  public.tipo_afiliacion  ta ON ta.id = e.tipo_afiliacion_id
                WHERE {where}
            ),
            totales AS (
                SELECT SUM(valor) AS gran_total FROM base
            )
            SELECT
                eps,
                COUNT(*)                                              AS total_eventos,
                COALESCE(SUM(valor), 0)                              AS total_facturado,
                COUNT(*) FILTER (WHERE tiene_contrato = TRUE)         AS eventos_con_contrato,
                COUNT(*) FILTER (WHERE tiene_contrato = FALSE)        AS eventos_sin_contrato,
                COUNT(*) FILTER (WHERE estado_id = {ESTADO_PENDIENTE}) AS pendientes,
                COUNT(*) FILTER (WHERE estado_id = {ESTADO_TERMINADO}) AS terminados,
                ROUND(
                    CASE WHEN (SELECT gran_total FROM totales) > 0
                         THEN COALESCE(SUM(valor), 0) * 100.0
                              / (SELECT gran_total FROM totales)
                         ELSE 0
                    END, 1
                )                                                     AS pct_participacion
            FROM base
            GROUP BY eps
            ORDER BY total_facturado DESC
            """,
            params
        )
        return [dict(r) for r in cur.fetchall()]


# ══════════════════════════════════════════════════════════════
# RESUMEN KPIs
# ══════════════════════════════════════════════════════════════

def obtener_resumen(entidad_id: int, rol: str = "ops", ops_id=None,
                    fecha_desde: str | None = None,
                    fecha_hasta: str | None = None) -> dict:
    """KPIs generales para las tarjetas del modulo."""
    oid = _ops_filtro(rol, ops_id)
    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        # Intentar RPC nativa
        try:
            cur.execute(
                """SELECT * FROM public.resumen_facturacion(
                       %s::integer, %s::integer, %s::date, %s::date
                   )""",
                (entidad_id, oid,
                 fecha_desde or None, fecha_hasta or None)
            )
            row = cur.fetchone()
            return dict(row) if row else {}
        except Exception:
            pass

        # Fallback directo
        conds, params = _conds_base(entidad_id, oid, fecha_desde, fecha_hasta)
        where = " AND ".join(conds)
        cur.execute(
            f"""SELECT
                    COUNT(*)                                      AS total_eventos,
                    COALESCE(SUM(valor), 0)                       AS total_facturado,
                    COUNT(*) FILTER (WHERE estado_id = {ESTADO_PENDIENTE})
                                                                  AS eventos_pendientes,
                    COUNT(*) FILTER (WHERE estado_id = {ESTADO_TERMINADO})
                                                                  AS eventos_terminados,
                    COUNT(*) FILTER (WHERE tiene_contrato = TRUE)  AS eventos_con_contrato,
                    COUNT(*) FILTER (WHERE tiene_contrato = FALSE
                                     AND eps_id IS NOT NULL)      AS eventos_sin_contrato
                FROM public.evento e
               WHERE {where}""",
            params
        )
        row = cur.fetchone()
        return dict(row) if row else {}


# ══════════════════════════════════════════════════════════════
# DATOS DE LA ENTIDAD
# ══════════════════════════════════════════════════════════════

def obtener_datos_entidad(entidad_id: int) -> dict:
    try:
        with Conexion(dict_cursor=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT nombre_entidad, nit, celular, correo "
                "FROM public.entidad WHERE id = %s",
                (entidad_id,)
            )
            row = cur.fetchone()
            return dict(row) if row else {}
    except Exception:
        return {}


# ══════════════════════════════════════════════════════════════
# HELPERS INTERNOS DE EXPORTACION
# ══════════════════════════════════════════════════════════════

_TITULOS_REPORTE = {
    "produccion":  "Reporte de Produccion General",
    "facturacion": "Reporte de Facturacion / Eventos Terminados",
    "cartera":     "Reporte de Cartera Operativa / Pendientes",
    "eps":         "Reporte Estrategico por EPS",
}

# Columnas por tipo de reporte (titulo, clave_dict, ancho_excel)
_COLS_PRODUCCION = [
    ("Nombre del paciente",  "nombre_paciente",       28),
    ("Tipo ID",              "tipo_identificacion",   12),
    ("N Identificacion",     "numero_identificacion", 18),
    ("EPS",                  "eps",                   22),
    ("Tipo afiliacion",      "tipo_afiliacion",       18),
    ("Motivo",               "motivo_evento",         36),
    ("N Admision",           "numero_admision",       14),
    ("Codigo evento",        "codigo_evento",         14),
    ("N Factura",            "numero_factura",        14),
    ("Valor ($)",            "valor",                 12),
    ("Estado",               "estado_evento",         12),
    ("Fecha",                "fecha_evento",          12),
    ("Registrado por",       "registrado_por",        18),
]

_COLS_FACTURACION = [
    ("Nombre del paciente",  "nombre_paciente",       28),
    ("Tipo ID",              "tipo_identificacion",   12),
    ("N Identificacion",     "numero_identificacion", 18),
    ("EPS",                  "eps",                   22),
    ("Tipo afiliacion",      "tipo_afiliacion",       18),
    ("N Admision",           "numero_admision",       14),
    ("Codigo evento",        "codigo_evento",         14),
    ("N Factura",            "numero_factura",        16),
    ("Valor ($)",            "valor",                 12),
    ("Fecha",                "fecha_evento",          12),
    ("Registrado por",       "registrado_por",        18),
]

_COLS_CARTERA = [
    ("Nombre del paciente",  "nombre_paciente",       28),
    ("Tipo ID",              "tipo_identificacion",   12),
    ("N Identificacion",     "numero_identificacion", 18),
    ("EPS",                  "eps",                   22),
    ("Tipo afiliacion",      "tipo_afiliacion",       18),
    ("Motivo",               "motivo_evento",         36),
    ("N Admision",           "numero_admision",       14),
    ("Dias transcurridos",   "dias_transcurridos",    14),
    ("Fecha evento",         "fecha_evento",          12),
    ("Registrado por",       "registrado_por",        18),
]

_COLS_EPS = [
    ("EPS",                  "eps",                   30),
    ("Total eventos",        "total_eventos",         14),
    ("Total facturado ($)",  "total_facturado",       16),
    ("Con contrato",         "eventos_con_contrato",  14),
    ("Sin contrato",         "eventos_sin_contrato",  14),
    ("Pendientes",           "pendientes",            12),
    ("Terminados",           "terminados",            12),
    ("% Participacion",      "pct_participacion",     14),
]

_COLS_MAP = {
    "produccion":  _COLS_PRODUCCION,
    "facturacion": _COLS_FACTURACION,
    "cartera":     _COLS_CARTERA,
    "eps":         _COLS_EPS,
}

# Columnas con valor monetario
_COLS_VALOR   = {"valor", "total_facturado"}
_COLS_CENTRO  = {"fecha_evento", "dias_transcurridos", "tipo_identificacion",
                 "numero_identificacion", "estado_evento"}
_COLS_PORCENT = {"pct_participacion"}


def _fmt_val(key: str, raw) -> str:
    """Formatea un valor para mostrar en tabla PDF."""
    if raw is None or raw == "":
        return "--"
    if key in _COLS_VALOR:
        return f"${float(raw):,.0f}"
    if key in _COLS_PORCENT:
        return f"{float(raw):.1f}%"
    if key == "fecha_evento":
        return str(raw)[:10]
    if key == "dias_transcurridos":
        return f"{int(raw)} dias"
    return str(raw)


# ══════════════════════════════════════════════════════════════
# EXPORTAR PDF
# ══════════════════════════════════════════════════════════════

def exportar_pdf(filas: list[dict], entidad: dict, ruta: str,
                 tipo_reporte: str = "produccion",
                 fecha_desde: str | None = None,
                 fecha_hasta: str | None = None,
                 ops_nombre: str | None = None) -> Resultado:
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.colors import HexColor

        C_ACCENT  = HexColor("#2D6ADF")
        C_DARK    = HexColor("#0D1117")
        C_BORDER  = HexColor("#30363D")
        C_ROW_ALT = HexColor("#F5F7FA")
        C_MUTED   = HexColor("#6B7280")
        C_TEXT    = HexColor("#1A1A2E")
        C_WHITE   = colors.white
        C_TOT_BG  = HexColor("#1C3A6E")

        PAGE = landscape(A4)
        doc = SimpleDocTemplate(
            ruta, pagesize=PAGE,
            leftMargin=1.2*cm, rightMargin=1.2*cm,
            topMargin=1.2*cm, bottomMargin=1.2*cm,
        )

        st_cel = ParagraphStyle(
            "cel", fontSize=7.5, leading=10,
            textColor=C_TEXT, fontName="Helvetica",
        )
        st_hdr = ParagraphStyle(
            "hdr", fontSize=8, leading=10,
            textColor=C_WHITE, fontName="Helvetica-Bold",
        )
        st_titulo = ParagraphStyle(
            "titulo", fontSize=16, leading=20,
            textColor=C_ACCENT, fontName="Helvetica-Bold", spaceAfter=2,
        )
        st_sub = ParagraphStyle(
            "sub", fontSize=9, leading=12,
            textColor=C_MUTED, fontName="Helvetica", spaceAfter=2,
        )
        st_firma_lbl = ParagraphStyle(
            "flbl", fontSize=8, leading=11,
            textColor=C_MUTED, fontName="Helvetica",
            alignment=TA_CENTER,
        )
        st_firma_val = ParagraphStyle(
            "fval", fontSize=10, leading=13,
            textColor=C_TEXT, fontName="Helvetica-Bold",
            alignment=TA_CENTER,
        )

        nombre_e = entidad.get("nombre_entidad", "")
        nit_e    = entidad.get("nit", "")
        cel_e    = entidad.get("celular", "")
        correo_e = entidad.get("correo", "")

        # Periodo texto
        if fecha_desde and fecha_hasta:
            periodo_txt = f"{fecha_desde}  a  {fecha_hasta}"
        elif fecha_desde:
            periodo_txt = f"Desde {fecha_desde}"
        elif fecha_hasta:
            periodo_txt = f"Hasta {fecha_hasta}"
        else:
            periodo_txt = "Todos los periodos"

        story = []

        # Encabezado
        hdr_data = [[
            Paragraph(f"<b>{nombre_e}</b>", st_titulo),
            Paragraph(
                f"NIT: {nit_e}<br/>Tel: {cel_e}<br/>{correo_e}",
                st_sub
            )
        ]]
        hdr_tbl = Table(hdr_data, colWidths=["65%", "35%"])
        hdr_tbl.setStyle(TableStyle([
            ("VALIGN",    (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",     (1, 0), (1, 0),   "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0),  1.5, C_ACCENT),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]))
        story.append(hdr_tbl); story.append(Spacer(1, 6))

        titulo_rep = _TITULOS_REPORTE.get(tipo_reporte, "Reporte")
        sub_parts = [
            f"<b>{titulo_rep}</b>",
            f"Periodo: {periodo_txt}",
            f"Generado: {date.today().isoformat()}",
        ]
        if ops_nombre:
            sub_parts.append(f"OPS: {ops_nombre}")
        story.append(Paragraph("   |   ".join(sub_parts), st_sub))
        story.append(Spacer(1, 8))

        # KPI de totales rapidos
        if tipo_reporte != "eps":
            total_val = sum(float(r.get("valor") or 0) for r in filas)
            story.append(Paragraph(
                f"Total facturado: <b>${total_val:,.0f}</b>"
                f"  |  Registros: <b>{len(filas)}</b>",
                ParagraphStyle("tot", fontSize=10, textColor=C_ACCENT,
                               fontName="Helvetica-Bold", spaceAfter=8)
            ))
        else:
            total_ev = sum(int(r.get("total_eventos") or 0) for r in filas)
            total_val = sum(float(r.get("total_facturado") or 0) for r in filas)
            story.append(Paragraph(
                f"Total EPS: <b>{len(filas)}</b>"
                f"  |  Total eventos: <b>{total_ev}</b>"
                f"  |  Total facturado: <b>${total_val:,.0f}</b>",
                ParagraphStyle("tot", fontSize=10, textColor=C_ACCENT,
                               fontName="Helvetica-Bold", spaceAfter=8)
            ))

        # Tabla de datos
        cols_def = _COLS_MAP.get(tipo_reporte, _COLS_PRODUCCION)
        ancho_total = 25.7 * cm   # landscape A4 usable
        col_widths = []
        suma_anchos = sum(c[2] for c in cols_def)
        for _, _, aw in cols_def:
            col_widths.append(aw / suma_anchos * ancho_total)

        tbl_data = [[Paragraph(c[0], st_hdr) for c in cols_def]]
        for fila in filas:
            tbl_data.append([
                Paragraph(_fmt_val(key, fila.get(key)), st_cel)
                for _, key, _ in cols_def
            ])

        # Fila de total (solo para no-EPS)
        if tipo_reporte != "eps":
            total_f = sum(float(r.get("valor") or 0) for r in filas)
            fila_tot = [""] * len(cols_def)
            fila_tot[0] = Paragraph("<b>TOTAL</b>", st_hdr)
            for i, (_, key, _) in enumerate(cols_def):
                if key in _COLS_VALOR:
                    fila_tot[i] = Paragraph(f"<b>${total_f:,.0f}</b>", st_hdr)
            tbl_data.append(fila_tot)
        else:
            tbl_data.append([
                Paragraph("<b>TOTAL</b>", st_hdr),
                Paragraph(f"<b>{total_ev}</b>", st_hdr),
                Paragraph(f"<b>${total_val:,.0f}</b>", st_hdr),
            ] + [Paragraph("", st_hdr)] * (len(cols_def) - 3))

        tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
        tbl_style = [
            ("BACKGROUND",    (0, 0),  (-1, 0),  C_ACCENT),
            ("TEXTCOLOR",     (0, 0),  (-1, 0),  C_WHITE),
            ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
            ("BACKGROUND",    (0, -1), (-1, -1), C_TOT_BG),
            ("TEXTCOLOR",     (0, -1), (-1, -1), C_WHITE),
            ("GRID",          (0, 0),  (-1, -1), 0.4, C_BORDER),
            ("LINEBELOW",     (0, 0),  (-1, 0),  1.2, C_WHITE),
            ("LINEABOVE",     (0, -1), (-1, -1), 1.2, C_ACCENT),
            ("TOPPADDING",    (0, 0),  (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0),  (-1, -1), 3),
            ("LEFTPADDING",   (0, 0),  (-1, -1), 4),
            ("RIGHTPADDING",  (0, 0),  (-1, -1), 4),
            ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ]
        for ri in range(1, len(tbl_data) - 1):
            bg = C_ROW_ALT if ri % 2 == 0 else C_WHITE
            tbl_style.append(("BACKGROUND", (0, ri), (-1, ri), bg))
        tbl.setStyle(TableStyle(tbl_style))
        story.append(tbl)

        # Bloque de firma
        story.append(Spacer(1, 18))
        story.append(HRFlowable(width="100%", thickness=1, color=C_BORDER))
        story.append(Spacer(1, 12))
        firma_data = [[
            [
                Paragraph("<b>Nombre / Entidad</b>", st_firma_lbl),
                Spacer(1, 4),
                Paragraph(nombre_e, st_firma_val),
                Spacer(1, 2),
                Paragraph(f"NIT: {nit_e}", st_firma_lbl),
            ],
            [
                Paragraph("<b>Contacto</b>", st_firma_lbl),
                Spacer(1, 4),
                Paragraph(f"Tel: {cel_e}", st_firma_val),
                Spacer(1, 2),
                Paragraph(correo_e, st_firma_lbl),
            ],
            [
                Paragraph("<b>Firma</b>", st_firma_lbl),
                Spacer(1, 3.5*cm),
                HRFlowable(width="80%", thickness=1, color=C_BORDER),
                Paragraph("Firma del profesional / responsable", st_firma_lbl),
            ],
        ]]
        firma_tbl = Table(firma_data, colWidths=["33%", "33%", "34%"])
        firma_tbl.setStyle(TableStyle([
            ("VALIGN",     (0, 0), (-1, -1), "TOP"),
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("BOX",        (2, 0), (2, 0),   0.5, C_BORDER),
            ("BACKGROUND", (2, 0), (2, 0),   HexColor("#F9FAFB")),
        ]))
        story.append(firma_tbl)
        doc.build(story)
        return Resultado(True, f"PDF generado: {Path(ruta).name}", {"ruta": ruta})

    except Exception as e:
        return Resultado(False, f"Error al generar PDF: {e}")


# ══════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════

def exportar_excel(filas: list[dict], entidad: dict, ruta: str,
                   tipo_reporte: str = "produccion",
                   fecha_desde: str | None = None,
                   fecha_hasta: str | None = None,
                   ops_nombre: str | None = None) -> Resultado:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        AZUL      = "2D6ADF"
        AZUL_OSC  = "1C3A6E"
        BLANCO    = "FFFFFF"
        GRIS_CLR  = "F5F7FA"
        VERDE     = "3FB950"

        def _fill(h):  return PatternFill("solid", fgColor=h)
        def _font(bold=False, color="1A1A2E", size=10, italic=False):
            return Font(bold=bold, color=color, size=size,
                        italic=italic, name="Arial")
        def _aln(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        def _brd():
            s = Side(border_style="thin", color="D1D5DB")
            return Border(left=s, right=s, top=s, bottom=s)

        wb = Workbook()
        ws = wb.active
        ws.title = _TITULOS_REPORTE.get(tipo_reporte, "Reporte")[:31]

        nombre_e = entidad.get("nombre_entidad", "")
        nit_e    = entidad.get("nit", "")
        cel_e    = entidad.get("celular", "")
        correo_e = entidad.get("correo", "")

        if fecha_desde and fecha_hasta:
            periodo = f"{fecha_desde}  a  {fecha_hasta}"
        elif fecha_desde:
            periodo = f"Desde {fecha_desde}"
        elif fecha_hasta:
            periodo = f"Hasta {fecha_hasta}"
        else:
            periodo = "Todos los periodos"

        cols_def = _COLS_MAP.get(tipo_reporte, _COLS_PRODUCCION)
        n_cols   = len(cols_def)
        col_end  = get_column_letter(n_cols)

        # Encabezado
        ws.merge_cells(f"A1:{col_end}1")
        ws["A1"] = nombre_e
        ws["A1"].font      = _font(bold=True, color=BLANCO, size=14)
        ws["A1"].fill      = _fill(AZUL)
        ws["A1"].alignment = _aln("left")
        ws.row_dimensions[1].height = 26

        titulo_rep = _TITULOS_REPORTE.get(tipo_reporte, "Reporte")
        info_parts = [
            f"Reporte: {titulo_rep}",
            f"NIT: {nit_e}",
            f"Tel: {cel_e}",
            correo_e,
            f"Periodo: {periodo}",
        ]
        if ops_nombre:
            info_parts.append(f"OPS: {ops_nombre}")
        ws.merge_cells(f"A2:{col_end}2")
        ws["A2"] = "  |  ".join(info_parts)
        ws["A2"].font      = _font(color="6B7280", size=9, italic=True)
        ws["A2"].fill      = _fill("F0F4FF")
        ws["A2"].alignment = _aln("left")
        ws.row_dimensions[2].height = 15

        ws.merge_cells(f"A3:{col_end}3")
        ws["A3"] = f"Generado: {date.today().isoformat()}"
        ws["A3"].font      = _font(color="6B7280", size=8, italic=True)
        ws["A3"].alignment = _aln("left")
        ws.row_dimensions[3].height = 12

        # Cabecera de tabla
        HDR = 5
        for col, (titulo, _, ancho) in enumerate(cols_def, 1):
            letra = get_column_letter(col)
            ws.column_dimensions[letra].width = ancho
            c = ws.cell(row=HDR, column=col, value=titulo)
            c.font      = _font(bold=True, color=BLANCO, size=10)
            c.fill      = _fill(AZUL)
            c.alignment = _aln("center")
            c.border    = _brd()
        ws.row_dimensions[HDR].height = 22

        # Filas de datos
        total_val = 0.0
        total_ev  = 0
        for i, fila in enumerate(filas):
            rn = HDR + 1 + i
            bg = GRIS_CLR if i % 2 == 0 else BLANCO
            ws.row_dimensions[rn].height = 18
            for col, (_, key, _) in enumerate(cols_def, 1):
                raw = fila.get(key)
                if key in _COLS_VALOR:
                    v = float(raw) if raw else 0.0
                    total_val += v
                    c = ws.cell(row=rn, column=col, value=v)
                    c.number_format = "#,##0"
                    c.alignment     = _aln("right")
                elif key == "total_eventos":
                    v = int(raw) if raw else 0
                    total_ev += v
                    c = ws.cell(row=rn, column=col, value=v)
                    c.alignment = _aln("center")
                elif key == "pct_participacion":
                    v = float(raw) if raw else 0.0
                    c = ws.cell(row=rn, column=col, value=v)
                    c.number_format = "0.0\"%\""
                    c.alignment     = _aln("center")
                elif key == "dias_transcurridos":
                    c = ws.cell(row=rn, column=col,
                                value=int(raw) if raw is not None else 0)
                    c.alignment = _aln("center")
                elif key == "fecha_evento":
                    c = ws.cell(row=rn, column=col, value=str(raw)[:10] if raw else "")
                    c.alignment = _aln("center")
                else:
                    val_str = str(raw) if raw else "--"
                    c = ws.cell(row=rn, column=col, value=val_str)
                    c.alignment = _aln("left", wrap=(key in ("motivo_evento",)))
                c.font   = _font(size=9)
                c.fill   = _fill(bg)
                c.border = _brd()

        # Fila de totales
        tot_rn = HDR + 1 + len(filas)
        ws.row_dimensions[tot_rn].height = 22
        ws.merge_cells(
            start_row=tot_rn, start_column=1,
            end_row=tot_rn, end_column=max(1, n_cols - 2)
        )
        c_lbl = ws.cell(row=tot_rn, column=1, value="TOTAL GENERAL")
        c_lbl.font      = _font(bold=True, color=BLANCO, size=10)
        c_lbl.fill      = _fill(AZUL_OSC)
        c_lbl.alignment = _aln("center")
        c_lbl.border    = _brd()

        # Columna de valor total o eventos totales
        for col, (_, key, _) in enumerate(cols_def, 1):
            if key in _COLS_VALOR:
                c = ws.cell(
                    row=tot_rn, column=col,
                    value=f"=SUM({get_column_letter(col)}{HDR+1}"
                          f":{get_column_letter(col)}{tot_rn-1})"
                )
                c.font         = _font(bold=True, color=BLANCO, size=10)
                c.fill         = _fill(AZUL_OSC)
                c.number_format = "#,##0"
                c.alignment    = _aln("right")
                c.border       = _brd()
            elif key == "total_eventos":
                c = ws.cell(row=tot_rn, column=col, value=total_ev)
                c.font      = _font(bold=True, color=BLANCO, size=10)
                c.fill      = _fill(AZUL_OSC)
                c.alignment = _aln("center")
                c.border    = _brd()

        # Bloque firma
        fr = tot_rn + 3
        ws.cell(row=fr, column=1,
                value="DATOS DEL PROFESIONAL / ENTIDAD").font = \
            _font(bold=True, color=AZUL, size=10)
        for off, (lbl, val) in enumerate([
            ("Nombre / Entidad:", nombre_e),
            ("NIT:",              nit_e),
            ("Telefono:",         cel_e),
            ("Correo:",           correo_e),
        ], 1):
            r = fr + off
            ws.row_dimensions[r].height = 16
            ws.cell(row=r, column=1, value=lbl).font = \
                _font(bold=True, color="6B7280", size=9)
            c = ws.cell(row=r, column=2, value=val)
            c.font   = _font(size=9)
            c.border = Border(
                bottom=Side(border_style="thin", color="D1D5DB")
            )

        firma_r = fr + 6
        ws.row_dimensions[firma_r].height = 14
        ws.cell(row=firma_r, column=1, value="FIRMA:").font = \
            _font(bold=True, color="6B7280", size=9)
        for r in range(firma_r + 1, firma_r + 7):
            ws.row_dimensions[r].height = 20
        ws.merge_cells(
            start_row=firma_r + 1, start_column=1,
            end_row=firma_r + 6, end_column=5
        )

        ws.freeze_panes = ws.cell(row=HDR + 1, column=1)
        wb.save(ruta)
        return Resultado(True, f"Excel generado: {Path(ruta).name}",
                         {"ruta": ruta})

    except Exception as e:
        return Resultado(False, f"Error al generar Excel: {e}")

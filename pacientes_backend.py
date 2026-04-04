# -*- coding: utf-8 -*-
# =============================================================================
# pacientes_backend.py  -  v3.0
# Modulo de Gestion de Pacientes -- Sistema SIGES
#
# CAMBIOS v3.0  (sobre v2.0)
#   . Arquitectura staging: Excel -> Prevalidador -> staging_paciente -> SQL masivo
#   . Prevalidador inteligente con autocorreccion y clasificacion por estado
#     (VALIDA / CORREGIDA / RECHAZADA / DUPLICADA-EN-ARCHIVO)
#   . Deduplicacion interna antes de tocar la BD
#   . Carga desde staging via INSERT...SELECT con JOIN en PostgreSQL
#     (sin bucles Python por fila, sin consultas por registro)
#   . Smart-update via ON CONFLICT + COALESCE (el motor SQL hace el trabajo)
#   . Reporte Excel: Resumen + Detalle + hoja "Rechazadas" lista para corregir
#   . Migracion idempotente de columna 'direccion' en staging_paciente
#
# FLUJO:
#   Excel -> _leer_archivo -> _prevalidar_lote -> staging_paciente
#         -> _commit_desde_staging (1 SQL masivo) -> reporte Excel
#
# ACCESO:
#   PERMITIDO -> rol='admin' | rol='ops'
#   Carga masiva disponible para admin y maestro.
#
# EJECUTOR:
#   { 'rol': 'admin'|'ops', 'ops_id': int|None,
#     'entidad_id': int, 'es_maestro': bool, 'nombre': str }
# =============================================================================

from __future__ import annotations

import csv
import datetime
import logging
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set, Tuple

from conexion import Conexion

logger = logging.getLogger("siges.pacientes")

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ──────────────────────────────────────────────────────────────────────────────

_CHUNK_EXCEL   = 20_000   # filas por chunk del prevalidador
_CHUNK_STAGING = 5_000    # filas por INSERT a staging
MAX_FILAS_LOTE = 10_000_000

_TIPOS_SIN_DOC: Set[str] = {"AS", "MS"}

_ALIAS_TIPO_DOC: Dict[str, str] = {
    "CEDULA":               "CC",
    "CEDULA DE CIUDADANIA": "CC",
    "CEDULA CIUDADANIA":    "CC",
    "TARJETA DE IDENTIDAD": "TI",
    "REGISTRO CIVIL":       "RC",
    "PASAPORTE":            "PP",
    "EXTRANJERIA":          "CE",
    "CEDULA EXTRANJERIA":   "CE",
    "PERMISO":              "PEP",
}

_COLOR: Dict[str, str] = {
    "azul_osc": "1E3A5F", "azul_med": "2D6ADF", "azul_cla": "EEF3FF",
    "rojo_osc": "9B1C1C", "rojo_cla": "FEF2F2",
    "ama_osc":  "92400E", "ama_cla":  "FFFBEB",
    "ver_osc":  "065F46", "ver_cla":  "ECFDF5",
    "gris": "F5F7FA", "blanco": "FFFFFF", "borde": "D1D5DB",
}

ESTADO_VALIDA    = "VALIDA"
ESTADO_CORREGIDA = "CORREGIDA"
ESTADO_RECHAZADA = "RECHAZADA"
ESTADO_DUPLICADA = "DUPLICADA"


# ──────────────────────────────────────────────────────────────────────────────
# DATACLASSES
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class Resultado:
    ok:      bool
    mensaje: str
    datos:   object = field(default=None)


@dataclass
class FilaValidada:
    """Fila despues del prevalidador, lista para staging o para reporte."""
    numero_fila:          int
    estado:               str
    tipo_documento:       str
    numero_documento:     str
    primer_apellido:      str
    segundo_apellido:     str
    primer_nombre:        str
    segundo_nombre:       Optional[str]
    fecha_nacimiento:     Optional[datetime.date]
    sexo:                 Optional[str]
    telefono:             Optional[str]
    eps_codigo:           Optional[str]
    tipo_afiliacion:      Optional[str]
    municipio_residencia: Optional[str]
    zona_residencia:      Optional[str]
    direccion:            Optional[str]
    correcciones:         List[str] = field(default_factory=list)
    errores:              List[str] = field(default_factory=list)

    @property
    def error_descripcion(self) -> Optional[str]:
        if self.estado == ESTADO_RECHAZADA:
            return " | ".join(self.errores)
        if self.estado == ESTADO_DUPLICADA:
            primera = self.errores[0] if self.errores else "?"
            return f"Duplicado en archivo (primer registro en fila {primera})"
        return None

    @property
    def cargable(self) -> bool:
        return self.estado in (ESTADO_VALIDA, ESTADO_CORREGIDA)


# ──────────────────────────────────────────────────────────────────────────────
# CONTROL DE ACCESO
# ──────────────────────────────────────────────────────────────────────────────

def puede_acceder(ejecutor: dict) -> bool:
    return ejecutor.get("rol", "") in ("admin", "ops")


def _check(ejecutor: dict) -> Optional[Resultado]:
    if not puede_acceder(ejecutor):
        return Resultado(False, "Acceso denegado. Inicia sesion para gestionar pacientes.")
    return None


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS DE PARSEO
# ──────────────────────────────────────────────────────────────────────────────

def _ops_none(v) -> Optional[int]:
    if v is None or v == 0 or str(v).strip() == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _sexo(raw) -> Optional[str]:
    s = str(raw or "").strip().upper()
    if s in ("M", "MASCULINO", "HOMBRE", "MALE", "H"):
        return "M"
    if s in ("F", "FEMENINO", "MUJER", "FEMALE"):
        return "F"
    if s in ("O", "OTRO", "OTHER"):
        return "O"
    return None


def _zona(raw) -> Optional[str]:
    z = str(raw or "").strip().upper()
    if z in ("URBANA", "U", "URBAN", "1"):
        return "Urbana"
    if z in ("RURAL", "R", "2"):
        return "Rural"
    return None


def _fecha(raw) -> Optional[datetime.date]:
    """String, serial Excel o date/datetime -> datetime.date. Nunca lanza."""
    if raw is None:
        return None
    if isinstance(raw, datetime.datetime):
        return raw.date()
    if isinstance(raw, datetime.date):
        return raw
    if isinstance(raw, (int, float)):
        n = int(raw)
        if 1 <= n <= 2_958_465:
            return datetime.date(1899, 12, 30) + datetime.timedelta(days=n)
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("none", "null", "nan", ""):
        return None
    for fmt in (
        "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y",
        "%d/%m/%y", "%Y/%m/%d", "%m/%d/%Y",
        "%d.%m.%Y", "%Y.%m.%d",
    ):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _limpiar_documento(doc: str) -> str:
    return doc.replace(" ", "").replace(".", "").replace("-", "").strip()


def _normalizar_fila(fila: dict) -> dict:
    return {
        (k.strip().rstrip(" *").lower() if k else ""): (
            str(v).strip() if v is not None else ""
        )
        for k, v in fila.items()
    }


def _g(fila_norm: dict, *claves: str) -> str:
    for clave in claves:
        val = fila_norm.get(clave.strip().rstrip(" *").lower(), "")
        if val:
            return str(val).strip()
    return ""


# ──────────────────────────────────────────────────────────────────────────────
# CATALOGOS
# ──────────────────────────────────────────────────────────────────────────────

def obtener_tipos_documento() -> List[dict]:
    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT id, abreviatura, nombre "
            "FROM public.tipo_documento WHERE activo = TRUE ORDER BY nombre"
        )
        return [dict(r) for r in cur.fetchall()]


def obtener_eps_activas(entidad_id: int) -> List[dict]:
    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT  e.id  AS eps_id,
                    COALESCE(e.codigo, '') AS codigo,
                    COALESCE(e.nombre, '') AS nombre,
                    public.eps_tiene_contrato(e.entidad_id, e.id, CURRENT_DATE) AS tiene_contrato
            FROM    public.eps e
            WHERE   e.entidad_id = %s AND e.activo = TRUE
            ORDER BY
                public.eps_tiene_contrato(e.entidad_id, e.id, CURRENT_DATE) DESC,
                e.nombre
            """,
            (entidad_id,)
        )
        return [dict(r) for r in cur.fetchall()]


def obtener_tipos_afiliacion() -> List[dict]:
    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, nombre, COALESCE(codigo,'') AS codigo
            FROM   public.tipo_afiliacion WHERE activo = TRUE
            ORDER BY
                CASE WHEN codigo IN ('01','02','03','04','05') THEN 0 ELSE 1 END,
                nombre
            """
        )
        return [dict(r) for r in cur.fetchall()]


# ──────────────────────────────────────────────────────────────────────────────
# LISTAR / BUSCAR
# ──────────────────────────────────────────────────────────────────────────────

def listar_pacientes(
    ejecutor:     dict,
    entidad_id:   int,
    filtro:       str  = "",
    solo_activos: bool = False,
    limite:       int  = 500,
    offset:       int  = 0,
) -> List[dict]:
    err = _check(ejecutor)
    if err:
        raise PermissionError(err.mensaje)
    like = f"%{filtro.strip()}%" if filtro.strip() else "%"
    cond = "AND p.activo = TRUE" if solo_activos else ""
    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT
                p.id                                                    AS paciente_id,
                p.entidad_id,
                td.abreviatura                                          AS tipo_doc,
                td.nombre                                               AS tipo_doc_nombre,
                p.numero_documento,
                p.primer_apellido,
                COALESCE(p.segundo_apellido,'')                         AS segundo_apellido,
                p.primer_nombre,
                COALESCE(p.segundo_nombre,'')                           AS segundo_nombre,
                CONCAT_WS(' ', p.primer_nombre, NULLIF(p.segundo_nombre,''),
                               p.primer_apellido, NULLIF(p.segundo_apellido,''))
                                                                        AS nombre_completo,
                p.fecha_nacimiento, p.sexo,
                COALESCE(p.municipio_residencia,'')                     AS municipio_residencia,
                COALESCE(p.zona_residencia,'')                          AS zona_residencia,
                COALESCE(p.direccion,'')                                AS direccion,
                COALESCE(p.telefono,'')                                 AS telefono,
                p.eps_id,
                COALESCE(ep.nombre,'')                                  AS eps_nombre,
                COALESCE(ep.codigo,'')                                  AS eps_codigo,
                p.tipo_afiliacion_id,
                COALESCE(ta.nombre,'')                                  AS tipo_afiliacion,
                p.activo, p.creado_en, p.actualizado_en, p.creado_por_ops,
                COALESCE(u.nombre_completo,'')                          AS creado_por_ops_nombre,
                COUNT(*) OVER()::int                                    AS total_count
            FROM   public.paciente         p
            JOIN   public.tipo_documento   td ON td.id = p.tipo_documento_id
            LEFT   JOIN public.eps          ep ON ep.id = p.eps_id
            LEFT   JOIN public.tipo_afiliacion ta ON ta.id = p.tipo_afiliacion_id
            LEFT   JOIN public.usuario_ops   u  ON u.id  = p.creado_por_ops
            WHERE  p.entidad_id = %s
              AND  (p.numero_documento ILIKE %s OR p.primer_apellido ILIKE %s
                    OR p.segundo_apellido ILIKE %s OR p.primer_nombre ILIKE %s
                    OR CONCAT_WS(' ', p.primer_nombre, p.segundo_nombre,
                                      p.primer_apellido, p.segundo_apellido) ILIKE %s)
              {cond}
            ORDER  BY p.primer_apellido, p.primer_nombre
            LIMIT  %s OFFSET %s
            """,
            (entidad_id, like, like, like, like, like, limite, offset)
        )
        return [dict(r) for r in cur.fetchall()]


def obtener_paciente(ejecutor: dict, entidad_id: int, paciente_id: int) -> Optional[dict]:
    err = _check(ejecutor)
    if err:
        raise PermissionError(err.mensaje)
    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT p.id AS paciente_id, p.entidad_id,
                   td.id AS tipo_doc_id, td.abreviatura AS tipo_doc, td.nombre AS tipo_doc_nombre,
                   p.numero_documento,
                   p.primer_apellido, COALESCE(p.segundo_apellido,'') AS segundo_apellido,
                   p.primer_nombre,   COALESCE(p.segundo_nombre,'')   AS segundo_nombre,
                   p.fecha_nacimiento, p.sexo,
                   COALESCE(p.municipio_residencia,'') AS municipio_residencia,
                   COALESCE(p.zona_residencia,'')      AS zona_residencia,
                   COALESCE(p.direccion,'')            AS direccion,
                   COALESCE(p.telefono,'')             AS telefono,
                   p.eps_id, COALESCE(ep.nombre,'') AS eps_nombre, COALESCE(ep.codigo,'') AS eps_codigo,
                   p.tipo_afiliacion_id, COALESCE(ta.nombre,'') AS tipo_afiliacion,
                   p.activo, p.creado_en, p.actualizado_en,
                   u.nombre_completo AS creado_por_ops_nombre
            FROM   public.paciente p
            JOIN   public.tipo_documento   td ON td.id  = p.tipo_documento_id
            LEFT   JOIN public.eps          ep ON ep.id = p.eps_id
            LEFT   JOIN public.tipo_afiliacion ta ON ta.id = p.tipo_afiliacion_id
            LEFT   JOIN public.usuario_ops   u  ON u.id  = p.creado_por_ops
            WHERE  p.id = %s AND p.entidad_id = %s
            """,
            (paciente_id, entidad_id)
        )
        row = cur.fetchone()
        return dict(row) if row else None


def buscar_pacientes_rapido(
    ejecutor: dict, entidad_id: int, texto: str, limite: int = 20
) -> List[dict]:
    err = _check(ejecutor)
    if err:
        return []
    try:
        with Conexion(dict_cursor=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT * FROM public.buscar_pacientes(%s, %s, %s)",
                (entidad_id, texto or None, limite)
            )
            return [dict(r) for r in cur.fetchall()]
    except Exception as e:
        logger.error("buscar_pacientes_rapido: %s", e)
        return []


# ──────────────────────────────────────────────────────────────────────────────
# GUARDAR INDIVIDUAL
# ──────────────────────────────────────────────────────────────────────────────

def guardar_paciente(
    ejecutor:    dict,
    entidad_id:  int,
    datos:       dict,
    paciente_id: Optional[int] = None,
) -> Resultado:
    """
    Crea o actualiza un paciente (formulario UI).
    Upsert por (entidad_id, numero_documento).
    Obligatorios: tipo_doc_abrev, numero_documento, primer_apellido, primer_nombre.
    """
    err = _check(ejecutor)
    if err:
        return err

    for campo in ("tipo_doc_abrev", "numero_documento", "primer_apellido", "primer_nombre"):
        if not str(datos.get(campo, "")).strip():
            return Resultado(False, f"El campo '{campo}' es obligatorio.")

    tipo_abrev = datos["tipo_doc_abrev"].strip().upper()
    num_doc    = _limpiar_documento(datos["numero_documento"])
    p_ape      = datos["primer_apellido"].strip().upper()
    p_nom      = datos["primer_nombre"].strip().upper()
    s_ape      = str(datos.get("segundo_apellido") or "").strip().upper()
    s_nom      = str(datos.get("segundo_nombre") or "").strip() or None
    direccion  = str(datos.get("direccion") or "").strip() or None
    telefono   = str(datos.get("telefono") or "").strip() or None
    municipio  = str(datos.get("municipio_residencia") or "").strip() or None
    zona       = _zona(datos.get("zona_residencia"))
    sexo       = _sexo(datos.get("sexo"))
    eps_id     = _ops_none(datos.get("eps_id"))
    afil_id    = _ops_none(datos.get("tipo_afiliacion_id"))
    fecha_nac  = _fecha(str(datos.get("fecha_nacimiento") or ""))
    ops_autor  = _ops_none(ejecutor.get("ops_id"))

    try:
        with Conexion(dict_cursor=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT id FROM public.tipo_documento WHERE abreviatura=%s AND activo=TRUE LIMIT 1",
                (tipo_abrev,)
            )
            td_row = cur.fetchone()
            if not td_row:
                return Resultado(False, f"Tipo de documento '{tipo_abrev}' no existe o no esta activo.")
            td_id = td_row["id"]

            if paciente_id is None:
                cur.execute(
                    "SELECT id FROM public.paciente WHERE entidad_id=%s AND numero_documento=%s LIMIT 1",
                    (entidad_id, num_doc)
                )
                row = cur.fetchone()
                if row:
                    paciente_id = row["id"]

            if paciente_id:
                cur.execute(
                    """
                    UPDATE public.paciente SET
                        tipo_documento_id=%s,
                        primer_apellido=%s, segundo_apellido=%s,
                        primer_nombre=%s,   segundo_nombre=%s,
                        fecha_nacimiento=%s, sexo=%s,
                        municipio_residencia=%s, zona_residencia=%s,
                        direccion=%s, telefono=%s,
                        eps_id=%s, tipo_afiliacion_id=%s
                    WHERE id=%s AND entidad_id=%s RETURNING id
                    """,
                    (td_id, p_ape, s_ape, p_nom, s_nom, fecha_nac, sexo,
                     municipio, zona, direccion, telefono,
                     eps_id, afil_id, paciente_id, entidad_id)
                )
                if not cur.fetchone():
                    return Resultado(False, "Paciente no encontrado en esta entidad.")
                return Resultado(True, "Paciente actualizado exitosamente.",
                                 {"paciente_id": paciente_id, "accion": "actualizado"})

            cur.execute(
                "SELECT id FROM public.paciente WHERE entidad_id=%s AND numero_documento=%s LIMIT 1",
                (entidad_id, num_doc)
            )
            if cur.fetchone():
                return Resultado(False, f"Ya existe un paciente con el documento '{num_doc}'.")

            cur.execute(
                """
                INSERT INTO public.paciente (
                    entidad_id, tipo_documento_id, numero_documento,
                    primer_apellido, segundo_apellido, primer_nombre, segundo_nombre,
                    fecha_nacimiento, sexo, municipio_residencia, zona_residencia,
                    direccion, telefono, eps_id, tipo_afiliacion_id, activo, creado_por_ops
                ) VALUES (%s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s, %s,%s,%s,%s, TRUE,%s) RETURNING id
                """,
                (entidad_id, td_id, num_doc,
                 p_ape, s_ape, p_nom, s_nom,
                 fecha_nac, sexo, municipio, zona,
                 direccion, telefono, eps_id, afil_id, ops_autor)
            )
            nuevo_id = cur.fetchone()["id"]
            return Resultado(True, "Paciente creado exitosamente.",
                             {"paciente_id": nuevo_id, "accion": "creado"})

    except Exception as e:
        err_str = str(e).lower()
        if "chk_paciente_sexo" in err_str:
            return Resultado(False, "Valor de Sexo invalido. Use M, F u O.")
        if "unique" in err_str:
            return Resultado(False, "Ya existe un paciente con ese documento.")
        logger.exception("guardar_paciente")
        return Resultado(False, f"Error al guardar: {str(e).split(chr(10))[0]}")


# ──────────────────────────────────────────────────────────────────────────────
# CAMBIAR ESTADO
# ──────────────────────────────────────────────────────────────────────────────

def cambiar_estado_paciente(
    ejecutor: dict, entidad_id: int, paciente_id: int, activo: bool
) -> Resultado:
    err = _check(ejecutor)
    if err:
        return err
    try:
        with Conexion(dict_cursor=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "UPDATE public.paciente SET activo=%s WHERE id=%s AND entidad_id=%s RETURNING id",
                (activo, paciente_id, entidad_id)
            )
            if not cur.fetchone():
                return Resultado(False, "Paciente no encontrado.")
        return Resultado(True, f"Paciente {'activado' if activo else 'desactivado'} correctamente.")
    except Exception as e:
        return Resultado(False, f"Error: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# ESTADISTICAS
# ──────────────────────────────────────────────────────────────────────────────

def stats_pacientes(ejecutor: dict, entidad_id: int) -> dict:
    err = _check(ejecutor)
    if err:
        return {"total": 0, "activos": 0, "inactivos": 0, "con_eps": 0, "sin_eps": 0}
    try:
        with Conexion(dict_cursor=True) as conn:
            cur = conn.cursor()
            cur.execute(
                """
                SELECT COUNT(*) AS total,
                       COUNT(*) FILTER (WHERE activo=TRUE)      AS activos,
                       COUNT(*) FILTER (WHERE activo=FALSE)     AS inactivos,
                       COUNT(*) FILTER (WHERE eps_id IS NOT NULL) AS con_eps,
                       COUNT(*) FILTER (WHERE eps_id IS NULL)   AS sin_eps
                FROM public.paciente WHERE entidad_id=%s
                """,
                (entidad_id,)
            )
            row = cur.fetchone()
            return {k: int(row[k]) for k in row.keys()}
    except Exception as e:
        logger.error("stats_pacientes: %s", e)
        return {"total": 0, "activos": 0, "inactivos": 0, "con_eps": 0, "sin_eps": 0}


# ──────────────────────────────────────────────────────────────────────────────
# PLANTILLA EXCEL
# ──────────────────────────────────────────────────────────────────────────────

COLUMNAS_PLANTILLA = [
    "Tipo_identificacion", "Numero_documento",
    "Primer_apellido",     "Segundo_apellido",
    "Primer_nombre",       "Segundo_nombre",
    "Fecha_nacimiento",    "Sexo",
    "Municipio_residencia","Zona_residencia",
    "Telefono",            "Direccion",
    "Codigo_EPS",          "Tipo_afiliacion",
]


def generar_plantilla_excel(ruta: str) -> Resultado:
    """Genera plantilla XLSX con instrucciones, ayudas, ejemplos y hoja de tipos."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        C = _COLOR
        def _fill(c):  return PatternFill("solid", fgColor=c)
        def _font(bold=False, color="1A1A2E", size=10, italic=False):
            return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
        def _aln(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        def _brd():
            s = Side(border_style="thin", color=C["borde"])
            return Border(left=s, right=s, top=s, bottom=s)

        wb = Workbook()
        ws = wb.active
        ws.title = "Carga Pacientes"

        ws.merge_cells("A1:N1")
        ws["A1"].value     = "SIGES -- Plantilla Carga Masiva de Pacientes"
        ws["A1"].font      = _font(bold=True, color=C["blanco"], size=13)
        ws["A1"].fill      = _fill(C["azul_med"])
        ws["A1"].alignment = _aln("left", "center")
        ws.row_dimensions[1].height = 28

        instrucciones = [
            "INSTRUCCIONES:",
            "Obligatorios (*): Tipo_identificacion, Primer_apellido, Primer_nombre. "
            "Numero_documento es obligatorio EXCEPTO para AS y MS (se genera automaticamente).",
            "Tipos validos: CC, TI, RC, CE, TE, NIT, PEP, PPT, CD, PP, TD, CNV, AS, MS. "
            "Tambien acepta: Cedula->CC, Tarjeta de Identidad->TI, Pasaporte->PP.",
            "Fecha_nacimiento: DD/MM/AAAA, AAAA-MM-DD o numero serial de Excel.",
            "Sexo: M o Masculino->M | F o Femenino->F | O->O | vacio->sin dato.",
            "Codigo_EPS y Tipo_afiliacion son OPCIONALES. El sistema los resuelve automaticamente.",
            "Duplicados en el archivo: solo se carga el primer registro; los demas van al reporte.",
        ]
        for i, txt in enumerate(instrucciones, 2):
            ws.merge_cells(f"A{i}:N{i}")
            c = ws[f"A{i}"]
            c.value     = txt
            c.font      = _font(size=9, italic=(i > 2), bold=(i == 2), color="374151")
            c.fill      = _fill(C["azul_cla"])
            c.alignment = _aln("left", "center", wrap=True)
            ws.row_dimensions[i].height = 14

        SEP = 9
        ws.merge_cells(f"A{SEP}:N{SEP}")
        ws[f"A{SEP}"].fill = _fill(C["azul_med"])
        ws.row_dimensions[SEP].height = 4

        COLS_DEF = [
            ("Tipo_identificacion *", 18), ("Numero_documento *", 20),
            ("Primer_apellido *",     20), ("Segundo_apellido",   20),
            ("Primer_nombre *",       20), ("Segundo_nombre",     18),
            ("Fecha_nacimiento",      18), ("Sexo",               10),
            ("Municipio_residencia",  22), ("Zona_residencia",    16),
            ("Telefono",              16), ("Direccion",          28),
            ("Codigo_EPS",            16), ("Tipo_afiliacion",    22),
        ]
        HDR = SEP + 1
        for col, (titulo, ancho) in enumerate(COLS_DEF, 1):
            ws.column_dimensions[get_column_letter(col)].width = ancho
            c = ws.cell(row=HDR, column=col, value=titulo)
            c.font      = _font(bold=True, color=C["blanco"], size=10)
            c.fill      = _fill(C["azul_med"])
            c.alignment = _aln("center", "center")
            c.border    = _brd()
        ws.row_dimensions[HDR].height = 22

        ayudas = [
            "CC, TI, AS, MS...", "Vacio si AS/MS", "Texto", "'' si no tiene",
            "Texto", "Opcional", "DD/MM/AAAA", "M/F/O",
            "Texto", "Urbana/Rural", "Solo numeros", "Texto",
            "CODIGO EPS (opcional)", "Nombre o codigo (opcional)",
        ]
        for col, txt in enumerate(ayudas, 1):
            c = ws.cell(row=HDR+1, column=col, value=txt)
            c.font      = _font(size=8, italic=True, color="6B7280")
            c.fill      = _fill(C["gris"])
            c.alignment = _aln("center", "center", wrap=True)
            c.border    = _brd()
        ws.row_dimensions[HDR+1].height = 16

        ejemplos = [
            ["CC","1234567890","PEREZ","GARCIA","JUAN","CARLOS","15/03/1985","M",
             "Bogota","Urbana","3001234567","Calle 10 No 5-20","EPS001","Contributivo"],
            ["TI","987654321","LOPEZ","","MARIA","","20/07/2005","F",
             "Medellin","Urbana","3109876543","","EPS002","Subsidiado"],
            ["AS","","DESCONOCIDO","","PACIENTE","","","",
             "Santa Marta","Urbana","","","",""],
            ["MS","","SIN DATOS","","MENOR","","","M",
             "Santa Marta","Rural","","","",""],
            ["CNV","","SIERRA","","JUAN","","","M",
             "Santa Marta","Rural","","","EPS003","Subsidiado"],
        ]
        for fi, fila in enumerate(ejemplos, HDR+2):
            for col, val in enumerate(fila, 1):
                c = ws.cell(row=fi, column=col, value=val)
                c.font      = _font(size=9, color="374151")
                c.fill      = _fill(C["blanco"] if fi % 2 == 0 else C["gris"])
                c.alignment = _aln("left", "center")
                c.border    = _brd()
            ws.row_dimensions[fi].height = 18
        ws.freeze_panes = ws.cell(row=HDR+2, column=1)

        ws2 = wb.create_sheet("Tipos Documento")
        ws2.merge_cells("A1:C1")
        ws2["A1"].value     = "Tipos de Documento Validos"
        ws2["A1"].font      = _font(bold=True, color=C["blanco"], size=11)
        ws2["A1"].fill      = _fill(C["azul_med"])
        ws2["A1"].alignment = _aln("center", "center")
        tipos_ref = [
            ("Abrev.", "Nombre completo", "Uso tipico"),
            ("RC","Registro Civil","Recien nacidos"),
            ("TI","Tarjeta de Identidad","Menores 7-17 anos"),
            ("CC","Cedula de Ciudadania","Adultos colombianos"),
            ("CE","Cedula de Extranjeria","Extranjeros residentes"),
            ("TE","Tarjeta de Extranjeria","Extranjeros menores"),
            ("NIT","NIT","Personas juridicas"),
            ("PP","Pasaporte","Colombianos o extranjeros"),
            ("PEP","Permiso Especial de Permanencia","Migrantes venezolanos"),
            ("PPT","Permiso de Proteccion Temporal","Migrantes venezolanos"),
            ("CD","Carne Diplomatico","Diplomaticos"),
            ("TD","Documento desconocido","Sin identificar"),
            ("CNV","Certificado de Nacido Vivo","Neonatos sin RC"),
            ("AS","Adulto sin Identificacion","Urgencias -- adulto sin ID"),
            ("MS","Menor sin Identificacion","Urgencias -- menor sin ID"),
        ]
        for ri, (a, b, cv) in enumerate(tipos_ref, 2):
            is_h = (ri == 2)
            for ci, val in enumerate([a, b, cv], 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.font      = _font(bold=is_h, color=C["blanco"] if is_h else "374151")
                cell.fill      = _fill(C["azul_med"] if is_h else (
                    C["azul_cla"] if ci == 1 else (C["gris"] if ri % 2 == 0 else C["blanco"])
                ))
                cell.alignment = _aln("center" if ci == 1 else "left", "center")
                cell.border    = _brd()
        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 24

        ruta_final = str(Path(ruta).with_suffix(".xlsx"))
        wb.save(ruta_final)
        return Resultado(True, f"Plantilla guardada en: {ruta_final}", {"ruta": ruta_final})

    except ImportError:
        return Resultado(False, "Instala openpyxl: pip install openpyxl")
    except Exception as e:
        return Resultado(False, f"Error al generar plantilla: {e}")


def generar_plantilla_csv(ruta: str) -> Resultado:
    return generar_plantilla_excel(ruta)


# ──────────────────────────────────────────────────────────────────────────────
# REPORTE EXCEL  (Resumen + Detalle + Rechazadas)
# ──────────────────────────────────────────────────────────────────────────────

def _generar_reporte_excel(
    ruta_reporte:   str,
    nombre_archivo: str,
    total:          int,
    creados:        int,
    actualizados:   int,
    sin_cambio:     int,
    rechazadas:     List[FilaValidada],
    advertencias:   List[Tuple[int, str, str]],
    duracion_seg:   float,
) -> Optional[str]:
    """
    Genera reporte Excel con tres hojas:
      Resumen     -- metricas con semaforo visual
      Detalle     -- errores/advertencias tabulados
      Rechazadas  -- filas originales listas para corregir y re-subir
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        C = _COLOR
        def _fill(c):  return PatternFill("solid", fgColor=c)
        def _font(bold=False, color="111827", size=10, italic=False):
            return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
        def _aln(h="left", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        def _brd(color=C["borde"]):
            s = Side(border_style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        wb        = Workbook()
        total_err = len(rechazadas)
        total_adv = len(advertencias)
        procesados = creados + actualizados + sin_cambio

        def _encabezado_hoja(ws, texto, col_fin="B"):
            ws.merge_cells(f"A1:{col_fin}1")
            ws["A1"].value     = texto
            ws["A1"].font      = _font(bold=True, color=C["blanco"], size=13)
            ws["A1"].fill      = _fill(C["azul_osc"])
            ws["A1"].alignment = _aln("center", "center")
            ws.row_dimensions[1].height = 30

        def _cabecera_tabla(ws, ri, cabeceras, color_fondo=None):
            color_fondo = color_fondo or C["azul_med"]
            for ci, cab in enumerate(cabeceras, 1):
                c = ws.cell(row=ri, column=ci, value=cab)
                c.font      = _font(bold=True, color=C["blanco"], size=10)
                c.fill      = _fill(color_fondo)
                c.alignment = _aln("center", "center")
                c.border    = _brd()
            ws.row_dimensions[ri].height = 20

        # == HOJA 1: RESUMEN ==================================================
        ws = wb.active
        ws.title = "Resumen"
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 16
        _encabezado_hoja(ws, "SIGES -- Reporte de Carga Masiva de Pacientes")

        ws.merge_cells("A2:B2")
        ws["A2"].value     = f"Archivo: {nombre_archivo}"
        ws["A2"].font      = _font(italic=True, color=C["blanco"], size=9)
        ws["A2"].fill      = _fill(C["azul_med"])
        ws["A2"].alignment = _aln("center", "center")
        ws.row_dimensions[2].height = 17

        ws.merge_cells("A3:B3")
        ws["A3"].value     = (
            f"Generado: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            f"   |   Duracion: {duracion_seg:.1f} s"
        )
        ws["A3"].font      = _font(italic=True, size=8, color="6B7280")
        ws["A3"].fill      = _fill(C["azul_cla"])
        ws["A3"].alignment = _aln("center", "center")
        ws.row_dimensions[3].height = 14

        metricas = [
            ("TOTAL DE FILAS EN ARCHIVO",            total,        C["azul_osc"], C["azul_cla"]),
            ("Filas cargadas (nuevas + actualizadas)", procesados,  C["azul_med"], C["blanco"]),
            ("  Pacientes NUEVOS creados",            creados,      C["ver_osc"],  C["ver_cla"]),
            ("  Pacientes ACTUALIZADOS",              actualizados, C["ver_osc"],  C["ver_cla"]),
            ("  Sin cambios (ya al dia)",              sin_cambio,   "4B5563",      C["gris"]),
            ("  Advertencias (EPS/afil./correcciones)",total_adv,   C["ama_osc"],  C["ama_cla"]),
            ("  RECHAZADAS (no cargadas)",            total_err,    C["rojo_osc"], C["rojo_cla"]),
        ]
        for ri, (etiq, valor, c_txt, c_fondo) in enumerate(metricas, 5):
            ws.row_dimensions[ri].height = 24
            for ci in (1, 2):
                cell = ws.cell(row=ri, column=ci)
                cell.fill      = _fill(c_fondo)
                cell.border    = _brd()
                cell.font      = _font(bold=(ri == 5), color=c_txt, size=10)
                cell.alignment = _aln("left" if ci == 1 else "center", "center")
            ws.cell(row=ri, column=1).value = etiq
            ws.cell(row=ri, column=2).value = valor

        if total_err > 0:
            nr = 5 + len(metricas) + 1
            ws.merge_cells(f"A{nr}:B{nr}")
            ws[f"A{nr}"].value     = "Vea las hojas 'Detalle' y 'Rechazadas' para corregir y re-subir."
            ws[f"A{nr}"].font      = _font(bold=True, color=C["rojo_osc"], size=9)
            ws[f"A{nr}"].fill      = _fill(C["rojo_cla"])
            ws[f"A{nr}"].alignment = _aln("center", "center")
            ws.row_dimensions[nr].height = 20

        # == HOJA 2: DETALLE ==================================================
        ws2 = wb.create_sheet("Detalle")
        _encabezado_hoja(ws2, "Detalle de Errores y Advertencias", col_fin="E")
        for ci, ancho in enumerate([10, 22, 26, 16, 62], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = ancho
        _cabecera_tabla(ws2, 2, ["N Fila", "Documento", "Campo", "Tipo", "Descripcion"])

        filas_det: List[Tuple] = []
        for fv in rechazadas:
            filas_det.append((fv.numero_fila, fv.numero_documento, "--", "Error",
                               fv.error_descripcion or "Rechazada"))
        for fila, campo, msg in advertencias:
            filas_det.append((fila, "--", campo, "Advertencia", msg))
        filas_det.sort(key=lambda r: (r[3] != "Error", r[0]))

        for ri, (fila, doc, campo, tipo, desc) in enumerate(filas_det, 3):
            es_err  = (tipo == "Error")
            c_t_txt = C["rojo_osc"] if es_err else C["ama_osc"]
            c_t_fnd = C["rojo_cla"] if es_err else C["ama_cla"]
            bg      = c_t_fnd if ri % 2 == 0 else C["blanco"]
            for ci, val in enumerate([fila, doc, campo, tipo, desc], 1):
                cell = ws2.cell(row=ri, column=ci, value=val)
                cell.border    = _brd()
                cell.alignment = _aln("center" if ci in (1, 4) else "left", "center", wrap=(ci == 5))
                cell.font      = _font(color=c_t_txt if ci == 4 else "374151", bold=(ci == 4), size=9)
                cell.fill      = _fill(c_t_fnd if ci == 4 else bg)
            ws2.row_dimensions[ri].height = 16

        if not filas_det:
            ws2.merge_cells("A3:E3")
            ws2["A3"].value     = "Sin errores ni advertencias -- carga perfecta."
            ws2["A3"].font      = _font(bold=True, color=C["ver_osc"], size=10)
            ws2["A3"].fill      = _fill(C["ver_cla"])
            ws2["A3"].alignment = _aln("center", "center")
            ws2.row_dimensions[3].height = 22
        ws2.freeze_panes = ws2["A3"]

        # == HOJA 3: RECHAZADAS (para corregir y re-subir) ====================
        ws3 = wb.create_sheet("Rechazadas")
        _encabezado_hoja(ws3, "Filas Rechazadas -- Corregir y Re-subir", col_fin="O")
        cols_r = [
            ("N Fila", 12), ("Motivo del rechazo", 42),
            ("Tipo_identificacion", 18), ("Numero_documento", 20),
            ("Primer_apellido", 20),  ("Segundo_apellido", 20),
            ("Primer_nombre", 20),    ("Segundo_nombre", 18),
            ("Fecha_nacimiento", 18), ("Sexo", 10),
            ("Municipio", 22),         ("Zona", 14),
            ("Telefono", 16),          ("Direccion", 28),
            ("Codigo_EPS", 16),
        ]
        for ci, (titulo, ancho) in enumerate(cols_r, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = ancho
        _cabecera_tabla(ws3, 2, [t for t, _ in cols_r], color_fondo=C["rojo_osc"])

        for ri, fv in enumerate(rechazadas, 3):
            bg = C["rojo_cla"] if ri % 2 == 0 else C["blanco"]
            vals = [
                fv.numero_fila,
                fv.error_descripcion or "",
                fv.tipo_documento,
                fv.numero_documento,
                fv.primer_apellido,
                fv.segundo_apellido,
                fv.primer_nombre,
                fv.segundo_nombre or "",
                fv.fecha_nacimiento.strftime("%d/%m/%Y") if fv.fecha_nacimiento else "",
                fv.sexo or "",
                fv.municipio_residencia or "",
                fv.zona_residencia or "",
                fv.telefono or "",
                fv.direccion or "",
                fv.eps_codigo or "",
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws3.cell(row=ri, column=ci, value=val)
                cell.border    = _brd()
                cell.alignment = _aln("left" if ci > 2 else "center", "center", wrap=(ci == 2))
                cell.font      = _font(size=9, color="374151")
                cell.fill      = _fill(C["rojo_cla"] if ci == 2 else bg)
            ws3.row_dimensions[ri].height = 16

        if not rechazadas:
            ws3.merge_cells("A3:O3")
            ws3["A3"].value     = "No hubo filas rechazadas."
            ws3["A3"].font      = _font(bold=True, color=C["ver_osc"], size=10)
            ws3["A3"].fill      = _fill(C["ver_cla"])
            ws3["A3"].alignment = _aln("center", "center")
            ws3.row_dimensions[3].height = 22
        ws3.freeze_panes = ws3["A3"]

        ruta_final = str(Path(ruta_reporte).with_suffix(".xlsx"))
        wb.save(ruta_final)
        return ruta_final

    except Exception as e:
        logger.warning("_generar_reporte_excel: %s", e)
        return None


# ──────────────────────────────────────────────────────────────────────────────
# CARGA MASIVA -- LECTURA DE ARCHIVO
# ──────────────────────────────────────────────────────────────────────────────

def _leer_archivo(ruta: Path) -> List[dict]:
    """Lee CSV o XLSX y retorna lista de dicts raw. XLSX: read_only streaming."""
    ext = ruta.suffix.lower()

    if ext == ".csv":
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                with open(ruta, newline="", encoding=enc) as f:
                    return list(csv.DictReader(f))
            except UnicodeDecodeError:
                continue
        raise ValueError("No se pudo leer el CSV. Guardelo como UTF-8.")

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Instala openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active

        _buscar = {
            "tipo_identificacion", "numero_documento",
            "primer_apellido", "primer_nombre",
            "codigo_eps", "tipo doc", "documento", "apellido",
        }
        hdr_row = 1
        for row in ws.iter_rows(min_row=1, max_row=20):
            vals = {str(c.value or "").strip().rstrip(" *").lower() for c in row if c.value}
            if vals & _buscar:
                hdr_row = row[0].row
                break

        enc = [
            str(c.value or "").strip().rstrip(" *") if c.value else None
            for c in ws[hdr_row]
        ]

        inicio_datos = hdr_row + 1
        _palabras_ayuda = {
            "solo numeros", "texto", "opcional", "dd/mm/aaaa",
            "m/f/o", "urbana/rural", "codigo eps", "nombre tipo",
            "vacio si as/ms", "codigo eps (opcional)",
        }
        try:
            fila_sig = list(ws.iter_rows(
                min_row=hdr_row+1, max_row=hdr_row+1, values_only=True
            ))[0]
            fd = dict(zip(enc, [str(v).strip() if v else "" for v in fila_sig]))
            num_doc_sig = next(
                (v for k, v in fd.items()
                 if k and k.strip().rstrip(" *").lower() == "numero_documento"),
                ""
            )
            if not num_doc_sig or num_doc_sig.lower() in _palabras_ayuda or len(num_doc_sig) > 40:
                inicio_datos = hdr_row + 2
        except (IndexError, StopIteration):
            pass

        filas = []
        for row in ws.iter_rows(min_row=inicio_datos, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            filas.append(dict(zip(enc, [v if v is not None else "" for v in row])))
        wb.close()
        return filas

    raise ValueError("Formato no soportado. Use .csv o .xlsx")


# ──────────────────────────────────────────────────────────────────────────────
# PREVALIDADOR INTELIGENTE
# ──────────────────────────────────────────────────────────────────────────────

def _prevalidar_lote(
    filas_raw:    List[dict],
    offset_fila:  int,
    mapa_td:      Dict[str, int],
    mapa_eps:     Dict[str, int],
    mapa_eps_nom: Dict[str, int],
    mapa_afil:    Dict[str, int],
    entidad_id:   int,
    seq_sinid:    List[int],
    vistos:       Dict[str, int],
) -> Tuple[List[FilaValidada], List[FilaValidada]]:
    """
    Valida y autocorrige un chunk de filas crudas en Python puro (O(n), sin BD).

    Estado posible por fila:
      VALIDA    -- todos los campos correctos, sin cambios
      CORREGIDA -- se cargara con correcciones automaticas
      RECHAZADA -- campo obligatorio invalido, no se carga
      DUPLICADA -- mismo documento ya aparecio antes en el archivo

    Retorna: (cargables, rechazadas)
    """
    cargables:  List[FilaValidada] = []
    rechazadas: List[FilaValidada] = []

    for i, fila_raw in enumerate(filas_raw):
        num_fila = offset_fila + i + 2
        f = _normalizar_fila(fila_raw)
        correcciones: List[str] = []
        errores:      List[str] = []

        # Tipo de identificacion
        tipo_raw  = _g(f, "tipo_identificacion", "tipo_doc", "tipo documento")
        tipo      = tipo_raw.strip().upper()
        if not tipo:
            errores.append("Tipo_identificacion: vacio")
        else:
            tipo_norm = tipo.replace(".", "").replace("-", "")
            if tipo_norm not in mapa_td:
                tipo_alt = _ALIAS_TIPO_DOC.get(tipo_norm)
                if tipo_alt and tipo_alt in mapa_td:
                    correcciones.append(f"Tipo_identificacion: '{tipo_raw}' -> '{tipo_alt}'")
                    tipo = tipo_alt
                else:
                    errores.append(f"Tipo_identificacion '{tipo_raw}' no reconocido")
                    tipo = tipo_raw

        # Numero de documento
        num_doc_raw = _g(f, "numero_documento", "documento", "num_documento", "cedula")
        num_doc = ""
        if tipo in _TIPOS_SIN_DOC:
            if num_doc_raw:
                num_doc = _limpiar_documento(num_doc_raw)
            else:
                seq_sinid[0] += 1
                num_doc = f"{tipo}-{entidad_id}-{seq_sinid[0]:06d}"
                correcciones.append(f"Numero_documento generado: {num_doc}")
        else:
            if not num_doc_raw:
                errores.append("Numero_documento: vacio")
            else:
                limpio = _limpiar_documento(num_doc_raw)
                if not limpio:
                    errores.append("Numero_documento invalido (solo puntos/guiones/espacios)")
                else:
                    if limpio != num_doc_raw.strip():
                        correcciones.append(f"Numero_documento limpiado: '{num_doc_raw}' -> '{limpio}'")
                    num_doc = limpio

        # Primer apellido
        p_ape_raw = _g(f, "primer_apellido", "apellido1", "apellido")
        p_ape = p_ape_raw.strip().upper()
        if not p_ape:
            errores.append("Primer_apellido: vacio")

        # Primer nombre
        p_nom_raw = _g(f, "primer_nombre", "nombre1", "nombre")
        p_nom = p_nom_raw.strip().upper()
        if not p_nom:
            errores.append("Primer_nombre: vacio")

        # RECHAZAR si hay errores bloqueantes
        if errores:
            rechazadas.append(FilaValidada(
                numero_fila=num_fila, estado=ESTADO_RECHAZADA,
                tipo_documento=tipo, numero_documento=num_doc or num_doc_raw or "",
                primer_apellido=p_ape, segundo_apellido="",
                primer_nombre=p_nom, segundo_nombre=None,
                fecha_nacimiento=None, sexo=None,
                telefono=None, eps_codigo=None, tipo_afiliacion=None,
                municipio_residencia=None, zona_residencia=None, direccion=None,
                correcciones=correcciones, errores=errores,
            ))
            continue

        # Deduplicacion interna
        if num_doc in vistos:
            rechazadas.append(FilaValidada(
                numero_fila=num_fila, estado=ESTADO_DUPLICADA,
                tipo_documento=tipo, numero_documento=num_doc,
                primer_apellido=p_ape, segundo_apellido="",
                primer_nombre=p_nom, segundo_nombre=None,
                fecha_nacimiento=None, sexo=None,
                telefono=None, eps_codigo=None, tipo_afiliacion=None,
                municipio_residencia=None, zona_residencia=None, direccion=None,
                correcciones=[], errores=[str(vistos[num_doc])],
            ))
            continue
        vistos[num_doc] = num_fila

        # Campos opcionales -- autocorregir, nunca bloquear
        s_ape = (_g(f, "segundo_apellido", "apellido2") or "").strip().upper()
        s_nom = _g(f, "segundo_nombre", "nombre2") or None

        fecha_raw = (
            _g(f, "fecha_nacimiento", "fechanacimiento", "fecha nacimiento")
            or fila_raw.get("Fecha_nacimiento") or fila_raw.get("fecha_nacimiento")
        )
        fecha_nac = _fecha(fecha_raw)
        if fecha_raw and not fecha_nac:
            correcciones.append(f"Fecha_nacimiento '{fecha_raw}' no reconocida -> NULL")

        sexo_raw = _g(f, "sexo", "genero", "genero")
        sexo     = _sexo(sexo_raw)
        if sexo_raw and sexo is None:
            correcciones.append(f"Sexo '{sexo_raw}' no reconocido -> NULL")

        zona_raw = _g(f, "zona_residencia", "zona")
        zona     = _zona(zona_raw)
        if zona_raw and zona is None:
            correcciones.append(f"Zona_residencia '{zona_raw}' no reconocida -> NULL")

        telefono  = _g(f, "telefono", "celular", "tel") or None
        municipio = _g(f, "municipio_residencia", "municipio") or None
        direccion = _g(f, "direccion", "direccion_residencia") or None

        # EPS -- buscar por codigo, fallback por nombre
        codigo_eps_raw = _g(f, "codigo_eps", "eps", "codigo eps", "codigoeps").upper()
        eps_codigo_out: Optional[str] = None
        if codigo_eps_raw:
            if codigo_eps_raw in mapa_eps:
                eps_codigo_out = codigo_eps_raw
            else:
                eps_id_nom = mapa_eps_nom.get(codigo_eps_raw.lower())
                if eps_id_nom:
                    codigo_real = next(
                        (c for c, eid in mapa_eps.items() if eid == eps_id_nom), None
                    )
                    if codigo_real:
                        eps_codigo_out = codigo_real
                        correcciones.append(f"EPS buscada por nombre -> codigo '{codigo_real}'")
                if eps_codigo_out is None:
                    correcciones.append(f"EPS '{codigo_eps_raw}' no encontrada -> sin EPS")

        # Tipo afiliacion
        afil_raw = _g(f, "tipo_afiliacion", "tipo_de_afiliacion",
                       "afiliacion", "tipo afiliacion").lower()
        afil_out: Optional[str] = None
        if afil_raw:
            if afil_raw in mapa_afil:
                afil_out = afil_raw
            else:
                correcciones.append(f"Tipo_afiliacion '{afil_raw}' no encontrado -> NULL")

        estado = ESTADO_CORREGIDA if correcciones else ESTADO_VALIDA
        cargables.append(FilaValidada(
            numero_fila=num_fila, estado=estado,
            tipo_documento=tipo, numero_documento=num_doc,
            primer_apellido=p_ape, segundo_apellido=s_ape,
            primer_nombre=p_nom, segundo_nombre=s_nom,
            fecha_nacimiento=fecha_nac, sexo=sexo,
            telefono=telefono, eps_codigo=eps_codigo_out,
            tipo_afiliacion=afil_out,
            municipio_residencia=municipio, zona_residencia=zona,
            direccion=direccion,
            correcciones=correcciones, errores=[],
        ))

    return cargables, rechazadas


# ──────────────────────────────────────────────────────────────────────────────
# MIGRACION STAGING (idempotente)
# ──────────────────────────────────────────────────────────────────────────────

def _asegurar_col_staging() -> None:
    """Agrega columna 'direccion' a staging_paciente si no existe."""
    try:
        with Conexion() as conn:
            cur = conn.cursor()
            cur.execute(
                """
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_schema='public'
                          AND table_name='staging_paciente'
                          AND column_name='direccion'
                    ) THEN
                        ALTER TABLE public.staging_paciente ADD COLUMN direccion text;
                    END IF;
                END $$;
                """
            )
    except Exception as e:
        logger.warning("_asegurar_col_staging: %s", e)


# ──────────────────────────────────────────────────────────────────────────────
# COMMIT DESDE STAGING (1 SQL masivo)
# ──────────────────────────────────────────────────────────────────────────────

def _commit_desde_staging(
    lote_id: int, entidad_id: int, ops_autor: Optional[int]
) -> Tuple[int, int]:
    """
    INSERT ... SELECT desde staging_paciente -> paciente.
    PostgreSQL resuelve tipo_documento_id, eps_id y tipo_afiliacion_id via JOIN.
    Sin bucle Python. Sin consultas por fila.
    Retorna (creados, actualizados).
    """
    SQL = """
        INSERT INTO public.paciente (
            entidad_id, tipo_documento_id, numero_documento,
            primer_apellido, segundo_apellido,
            primer_nombre,   segundo_nombre,
            fecha_nacimiento, sexo,
            zona_residencia, municipio_residencia,
            telefono, direccion,
            eps_id, tipo_afiliacion_id,
            activo, creado_por_ops
        )
        SELECT
            %(eid)s,
            td.id,
            sp.numero_documento,
            sp.primer_apellido,
            COALESCE(sp.segundo_apellido, ''),
            sp.primer_nombre,
            NULLIF(sp.segundo_nombre, ''),
            sp.fecha_nacimiento,
            sp.sexo,
            sp.zona_residencia,
            sp.municipio_residencia,
            NULLIF(sp.telefono, ''),
            NULLIF(sp.direccion, ''),
            e.id,
            ta.id,
            TRUE,
            %(ops)s
        FROM public.staging_paciente sp
        JOIN public.tipo_documento td
            ON td.abreviatura = sp.tipo_documento AND td.activo = TRUE
        LEFT JOIN public.eps e
            ON e.entidad_id = %(eid)s
           AND e.codigo     = sp.eps_codigo
           AND e.activo     = TRUE
        LEFT JOIN public.tipo_afiliacion ta
            ON LOWER(ta.nombre) = LOWER(sp.tipo_afiliacion)
           AND ta.activo = TRUE
        WHERE sp.lote_id = %(lid)s
          AND sp.error_descripcion IS NULL
        ON CONFLICT ON CONSTRAINT uq_paciente_doc
        DO UPDATE SET
            tipo_documento_id    = EXCLUDED.tipo_documento_id,
            primer_apellido      = EXCLUDED.primer_apellido,
            segundo_apellido     = EXCLUDED.segundo_apellido,
            primer_nombre        = EXCLUDED.primer_nombre,
            segundo_nombre       = COALESCE(EXCLUDED.segundo_nombre,       public.paciente.segundo_nombre),
            fecha_nacimiento     = COALESCE(EXCLUDED.fecha_nacimiento,     public.paciente.fecha_nacimiento),
            sexo                 = COALESCE(EXCLUDED.sexo,                 public.paciente.sexo),
            zona_residencia      = COALESCE(EXCLUDED.zona_residencia,      public.paciente.zona_residencia),
            municipio_residencia = COALESCE(EXCLUDED.municipio_residencia, public.paciente.municipio_residencia),
            telefono             = COALESCE(EXCLUDED.telefono,             public.paciente.telefono),
            direccion            = COALESCE(EXCLUDED.direccion,            public.paciente.direccion),
            eps_id               = COALESCE(EXCLUDED.eps_id,               public.paciente.eps_id),
            tipo_afiliacion_id   = COALESCE(EXCLUDED.tipo_afiliacion_id,   public.paciente.tipo_afiliacion_id)
        RETURNING (xmax = 0) AS es_nuevo
    """
    creados = actualizados = 0
    with Conexion(dict_cursor=True) as conn:
        cur = conn.cursor()
        cur.execute(SQL, {"eid": entidad_id, "ops": ops_autor, "lid": lote_id})
        for row in cur.fetchall():
            if row["es_nuevo"]:
                creados += 1
            else:
                actualizados += 1
    return creados, actualizados


# ──────────────────────────────────────────────────────────────────────────────
# CARGA MASIVA -- FUNCION PRINCIPAL  v3.0
# ──────────────────────────────────────────────────────────────────────────────

def procesar_carga_masiva(
    ejecutor:     dict,
    entidad_id:   int,
    ruta_archivo: str,
    on_progreso:  Optional[Callable[[int, int], None]] = None,
    ruta_reporte: Optional[str] = None,
) -> Resultado:
    """
    Carga masiva ultra-rapida -- arquitectura staging.

    Flujo:
      1. Leer archivo (XLSX streaming / CSV multi-encoding).
      2. 3 queries para cargar catalogos (tipo_doc, eps, afiliacion).
      3. Registrar lote.
      4. Prevalidar en Python (chunks, sin BD):
           autocorrecciones + clasificar VALIDA/CORREGIDA/RECHAZADA/DUPLICADA.
      5. INSERT masivo de validas+corregidas -> staging_paciente (execute_values).
      6. 1 INSERT...SELECT masivo: staging -> paciente (PostgreSQL resuelve JOINs).
      7. DELETE staging del lote.
      8. Cerrar lote.
      9. Reporte Excel (Resumen + Detalle + Rechazadas para re-subir).

    Args:
        ruta_reporte: Ruta del Excel de reporte. Si None, se genera al lado
                      del archivo con sufijo '_reporte_errores.xlsx'.

    Returns:
        Resultado.datos:
            total, creados, actualizados, sin_cambio,
            rechazadas (count), advertencias_count,
            errores (list[dict]), lote_id, ruta_reporte, duracion_seg
    """
    t0 = time.monotonic()

    err = _check(ejecutor)
    if err:
        return err

    ruta = Path(ruta_archivo)
    if not ruta.exists():
        return Resultado(False, "El archivo no existe.")

    # 1. Leer
    try:
        filas_raw = _leer_archivo(ruta)
    except ImportError as e:
        return Resultado(False, str(e))
    except Exception as e:
        return Resultado(False, f"Error al leer el archivo: {e}")

    if not filas_raw:
        return Resultado(False, "El archivo esta vacio o no tiene datos validos.")

    total = len(filas_raw)
    if on_progreso:
        on_progreso(0, total)

    # 2. Catalogos
    mapa_td:      Dict[str, int] = {}
    mapa_eps:     Dict[str, int] = {}
    mapa_eps_nom: Dict[str, int] = {}
    mapa_afil:    Dict[str, int] = {}
    ops_autor = _ops_none(ejecutor.get("ops_id"))

    try:
        with Conexion(dict_cursor=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT id, UPPER(TRIM(abreviatura)) AS a "
                "FROM public.tipo_documento WHERE activo=TRUE"
            )
            mapa_td = {r["a"]: r["id"] for r in cur.fetchall()}

            cur.execute(
                "SELECT id, UPPER(TRIM(codigo)) AS c, LOWER(TRIM(nombre)) AS n "
                "FROM public.eps "
                "WHERE entidad_id=%s AND activo=TRUE "
                "AND codigo IS NOT NULL AND TRIM(codigo)<>''",
                (entidad_id,)
            )
            for r in cur.fetchall():
                mapa_eps[r["c"]] = r["id"]
                if r["n"]:
                    mapa_eps_nom[r["n"]] = r["id"]

            cur.execute(
                "SELECT id, LOWER(TRIM(nombre)) AS n, "
                "COALESCE(LOWER(TRIM(codigo)),'') AS c "
                "FROM public.tipo_afiliacion WHERE activo=TRUE"
            )
            for r in cur.fetchall():
                mapa_afil[r["n"]] = r["id"]
                if r["c"]:
                    mapa_afil[r["c"]] = r["id"]
    except Exception as e:
        return Resultado(False, f"Error al cargar catalogos: {e}")

    # 3. Registrar lote
    lote_id: Optional[int] = None
    try:
        with Conexion(dict_cursor=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO public.carga_masiva_lote "
                "(entidad_id, tipo, nombre_archivo, total_filas, estado) "
                "VALUES (%s,'pacientes',%s,%s,'procesando') RETURNING id",
                (entidad_id, ruta.name, total)
            )
            lote_id = cur.fetchone()["id"]
    except Exception as e:
        logger.warning("Registro de lote fallido: %s", e)

    # 4. Prevalidar en chunks
    _asegurar_col_staging()

    todas_cargables:  List[FilaValidada] = []
    todas_rechazadas: List[FilaValidada] = []
    advertencias:     List[Tuple[int, str, str]] = []
    seq_sinid: List[int] = [0]
    vistos:    Dict[str, int] = {}

    for chunk_start in range(0, total, _CHUNK_EXCEL):
        chunk = filas_raw[chunk_start: chunk_start + _CHUNK_EXCEL]
        carg, rech = _prevalidar_lote(
            filas_raw=chunk, offset_fila=chunk_start,
            mapa_td=mapa_td, mapa_eps=mapa_eps, mapa_eps_nom=mapa_eps_nom,
            mapa_afil=mapa_afil, entidad_id=entidad_id,
            seq_sinid=seq_sinid, vistos=vistos,
        )
        todas_cargables.extend(carg)
        todas_rechazadas.extend(rech)
        if on_progreso:
            on_progreso(min(chunk_start + _CHUNK_EXCEL, total), total)

    for fv in todas_cargables:
        for corr in fv.correcciones:
            advertencias.append((fv.numero_fila, "autocorreccion", corr))

    if on_progreso:
        on_progreso(total, total)

    # 5. INSERT masivo a staging
    try:
        from psycopg2.extras import execute_values
    except ImportError:
        return Resultado(False, "Requiere psycopg2: pip install psycopg2-binary")

    STAGING_SQL = """
        INSERT INTO public.staging_paciente (
            lote_id, numero_fila,
            tipo_documento, numero_documento,
            primer_apellido, segundo_apellido,
            primer_nombre,   segundo_nombre,
            fecha_nacimiento, sexo,
            telefono, eps_codigo, tipo_afiliacion,
            municipio_residencia, zona_residencia,
            direccion, error_descripcion
        ) VALUES %s
    """

    def _a_staging(fv: FilaValidada, lid: Optional[int]) -> Tuple:
        return (
            lid, fv.numero_fila,
            fv.tipo_documento, fv.numero_documento,
            fv.primer_apellido, fv.segundo_apellido or "",
            fv.primer_nombre, fv.segundo_nombre,
            fv.fecha_nacimiento, fv.sexo,
            fv.telefono, fv.eps_codigo, fv.tipo_afiliacion,
            fv.municipio_residencia, fv.zona_residencia,
            fv.direccion,
            None,   # error_descripcion=NULL -> cargable
        )

    if todas_cargables:
        try:
            with Conexion() as conn:
                cur = conn.cursor()
                if lote_id:
                    cur.execute(
                        "DELETE FROM public.staging_paciente WHERE lote_id=%s", (lote_id,)
                    )
                for i in range(0, len(todas_cargables), _CHUNK_STAGING):
                    chunk = todas_cargables[i: i + _CHUNK_STAGING]
                    execute_values(cur, STAGING_SQL,
                                   [_a_staging(fv, lote_id) for fv in chunk],
                                   page_size=_CHUNK_STAGING)
        except Exception as e:
            logger.error("insert staging: %s", e)
            return Resultado(False, f"Error al escribir en staging: {str(e).split(chr(10))[0]}")

    # 6. Commit masivo staging -> paciente
    creados = actualizados = 0
    if todas_cargables and lote_id:
        try:
            creados, actualizados = _commit_desde_staging(lote_id, entidad_id, ops_autor)
        except Exception as e:
            logger.error("commit_desde_staging: %s", e)
            return Resultado(False, f"Error al mover datos a paciente: {str(e).split(chr(10))[0]}")

    sin_cambio = max(0, len(todas_cargables) - creados - actualizados)

    # 7. Limpiar staging
    if lote_id:
        try:
            with Conexion() as conn:
                conn.cursor().execute(
                    "DELETE FROM public.staging_paciente WHERE lote_id=%s", (lote_id,)
                )
        except Exception as e:
            logger.warning("limpiar staging: %s", e)

    # 8. Cerrar lote
    if lote_id:
        try:
            with Conexion() as conn:
                conn.cursor().execute(
                    "UPDATE public.carga_masiva_lote "
                    "SET filas_ok=%s, filas_error=%s, "
                    "estado='completado', completado_en=NOW() WHERE id=%s",
                    (creados + actualizados, len(todas_rechazadas), lote_id)
                )
        except Exception:
            pass

    # 9. Reporte Excel
    duracion = time.monotonic() - t0
    if ruta_reporte is None:
        ruta_reporte = str(ruta.parent / f"{ruta.stem}_reporte_errores.xlsx")

    ruta_rep = _generar_reporte_excel(
        ruta_reporte=ruta_reporte, nombre_archivo=ruta.name,
        total=total, creados=creados, actualizados=actualizados,
        sin_cambio=sin_cambio, rechazadas=todas_rechazadas,
        advertencias=advertencias, duracion_seg=duracion,
    )

    # 10. Resultado
    return Resultado(
        ok=True,
        mensaje=(
            f"Carga completada en {duracion:.1f}s: "
            f"{creados} nuevos, {actualizados} actualizados, "
            f"{sin_cambio} sin cambios, "
            f"{len(todas_rechazadas)} rechazadas de {total} filas."
        ),
        datos={
            "total":              total,
            "creados":            creados,
            "actualizados":       actualizados,
            "sin_cambio":         sin_cambio,
            "rechazadas":         len(todas_rechazadas),
            "advertencias_count": len(advertencias),
            "errores": [
                {
                    "fila":      fv.numero_fila,
                    "documento": fv.numero_documento,
                    "estado":    fv.estado,
                    "motivo":    fv.error_descripcion or "",
                }
                for fv in todas_rechazadas
            ],
            "lote_id":      lote_id,
            "ruta_reporte": ruta_rep,
            "duracion_seg": round(duracion, 2),
        },
    )